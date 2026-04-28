"""
Роутер для админского доступа к товарам (JWT авторизация)
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Header, Form, Body
from fastapi.responses import FileResponse, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, update, func, and_, or_, true
from pydantic import BaseModel, Field, field_validator
from typing import Optional, List, Dict, Any, Tuple
from decimal import Decimal

from api.products.database.database import get_session
from api.products.models.item import Item
from api.products.models.item_photo import ItemPhoto
from api.products.models.item_stock import ItemStock
from api.products.models.item_reservation import ItemReservation
from api.products.models.order_delivery import OrderDelivery
from api.products.models.delivery_status import DeliveryStatus
from api.products.models.item_group import ItemGroup
from api.products.models.item_type import ItemType
from api.products.models.size_chart import SizeChart
from api.shared.jwt_admin_deps import require_jwt_permission
from api.shared import bot_api_client as _tg_bridge
from api.shared.tg_notify_copy import (
    MINIAPP_ORDERS_HINT,
    MINIAPP_REVIEW_HINT,
    REVIEW_CHANNEL_DUPLICATES_HINT,
    SUPPORT_VIA_CHANNEL_HINT,
)
from api.products.utils.parcel import (
    DEFAULT_WEIGHT_KG,
    aggregate_parcel_dimensions,
    build_line_items_for_parcel,
)
from api.products.utils.promo_apply import delete_promo_redemptions_for_order
from api.shared.timezone import get_week_start_vladivostok
from api.products.schemas.item import ItemResponse, ItemPhotoResponse
from api.products.schemas.size_chart import (
    SizeChartResponse,
    SizeChartListItem,
    CreateSizeChartRequest,
    UpdateSizeChartRequest,
)
from api.products.schemas.item_group import (
    ItemGroupResponse, ItemGroupWithItemsResponse, CreateItemGroupRequest,
    UpdateItemGroupRequest, AddItemsToGroupRequest, RemoveItemFromGroupRequest
)
from api.products.routers.feed import invalidate_catalog_cache
from api.products.utils.feed_like_counts import get_feed_like_dislike_counts_map, like_dislike_for
from api.products.utils.item_type_count import change_item_type_count
from sqlalchemy.orm import joinedload
from sqlalchemy.orm.attributes import flag_modified
import httpx
import base64
import os
from pathlib import Path
from os import getenv
from pathlib import Path
import aiofiles
import uuid
import re
import logging
from PIL import Image
import io

logger = logging.getLogger(__name__)

router = APIRouter()

UPLOAD_DIR = Path("uploads/items")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

FINANCE_SERVICE_URL = getenv("FINANCE_SERVICE_URL", "http://finance-service:8003")
DELIVERY_SERVICE_URL = getenv("DELIVERY_SERVICE_URL", "http://delivery-service:8005")
PHOTO_GROUP_ID = getenv("PHOTO_GROUP_ID")
USERS_SERVICE_URL = getenv("USERS_SERVICE_URL", "http://users-service:8001")
INTERNAL_TOKEN = getenv("INTERNAL_TOKEN", "internal-secret-token-change-in-production")
TELEGRAM_BOT_USERNAME = (getenv("TELEGRAM_BOT_USERNAME") or getenv("BOT_USERNAME") or "").strip().lstrip("@")

MAX_IMAGE_DIMENSION = 1600

# Относительный путь фото кастомной позиции из upload-custom-photo (без path traversal)
_MANUAL_CUSTOM_PHOTO_PATH = re.compile(r"^uploads/items/[A-Za-z0-9._-]+\.jpg$", re.I)


def _is_safe_manual_custom_photo_relative_path(path: Optional[str]) -> bool:
    if not path or not isinstance(path, str):
        return False
    p = path.strip().replace("\\", "/")
    if ".." in p or p.startswith("/"):
        return False
    return bool(_MANUAL_CUSTOM_PHOTO_PATH.match(p))


async def save_compressed_image(upload: UploadFile) -> Path:
  """
  Сохранить изображение в UPLOAD_DIR сжатым для веба.
  - конвертация в JPEG
  - максимальная сторона не более MAX_IMAGE_DIMENSION
  - качество ~80, optimize=True
  При ошибке Pillow сохраняем оригинальный файл без изменений.
  """
  file_name = f"{uuid.uuid4()}.jpg"
  file_path = UPLOAD_DIR / file_name

  content = await upload.read()
  try:
      image = Image.open(io.BytesIO(content))
      if image.mode not in ("RGB", "L"):
          image = image.convert("RGB")
      width, height = image.size
      max_side = max(width, height)
      if max_side > MAX_IMAGE_DIMENSION:
          scale = MAX_IMAGE_DIMENSION / max_side
          new_size = (int(width * scale), int(height * scale))
          image = image.resize(new_size, Image.LANCZOS)
      image.save(file_path, format="JPEG", optimize=True, quality=82)
  except Exception:
      # Если не удалось обработать через Pillow — сохраняем как есть
      async with aiofiles.open(file_path, "wb") as f:
          await f.write(content)

  return file_path


async def send_photo_to_telegram_group_and_get_file_id(photo_id: int, file_path: str) -> Optional[str]:
    """Отправить фото в группу Telegram для получения file_id. Возвращает file_id или None"""
    if not _tg_bridge.telegram_outbound_configured() or not PHOTO_GROUP_ID:
        logger.warning("Исходящий Telegram или PHOTO_GROUP_ID не настроены, пропускаем отправку фото")
        return None
    
    # Проверяем, есть ли уже file_id у этого фото (чтобы не отправлять повторно)
    from api.products.database.database import async_session_maker
    from sqlalchemy import select
    async with async_session_maker() as check_session:
        result = await check_session.execute(
            select(ItemPhoto).where(ItemPhoto.id == photo_id)
        )
        existing_photo = result.scalar_one_or_none()
        if existing_photo and existing_photo.telegram_file_id:
            logger.info(f"Фото #{photo_id} уже имеет file_id ({existing_photo.telegram_file_id[:20]}...), пропускаем отправку")
            return existing_photo.telegram_file_id
        logger.info(f"Отправляю фото #{photo_id} в Telegram группу {PHOTO_GROUP_ID}, файл: {file_path}")
    
    try:
        # Проверяем существование файла
        photo_path = Path(file_path)
        if not photo_path.exists():
            # Пробуем разные варианты путей
            possible_paths = [
                Path(f"/app/{file_path}"),
                Path(file_path),
                Path(f"uploads/items/{photo_path.name}"),
                Path(f"/app/uploads/items/{photo_path.name}")
            ]
            for path in possible_paths:
                if path.exists():
                    photo_path = path
                    break
            else:
                logger.error(f"Файл не найден: {file_path}")
                return
        
        async with aiofiles.open(photo_path, "rb") as photo_file:
            photo_content = await photo_file.read()

            result = await _tg_bridge.telegram_send_photo_multipart(
                PHOTO_GROUP_ID,
                photo_path.name,
                photo_content,
                "image/jpeg",
                caption=f"photo_id:{photo_id}",
                timeout=30.0,
            )

            if result.get("ok") and result.get("result", {}).get("photo"):
                # Получаем file_id из самого большого фото
                photos = result["result"]["photo"]
                largest_photo = max(photos, key=lambda p: p.get("file_size", 0))
                file_id = largest_photo.get("file_id")

                if file_id:
                    logger.info(f"✅ Получен file_id для фото #{photo_id}: {file_id[:30]}...")
                    return file_id
                logger.warning(f"⚠️ File_id не найден в ответе Telegram для фото #{photo_id}")
                return None
            logger.error(f"❌ Telegram API вернул ошибку для фото #{photo_id}: {result}")
            return None
    except Exception as e:
        logger.error(f"Ошибка при отправке фото #{photo_id} в Telegram: {e}", exc_info=True)
        return None


async def upload_file_to_telegram_and_get_file_id(file: "UploadFile") -> Optional[str]:
    """Загрузить файл в Telegram группу и получить file_id для использования в постах."""
    if not _tg_bridge.telegram_outbound_configured() or not PHOTO_GROUP_ID:
        return None
    try:
        content = await file.read()
        filename = file.filename or "photo.jpg"
        ct = file.content_type or "image/jpeg"
        result = await _tg_bridge.telegram_send_photo_multipart(
            PHOTO_GROUP_ID,
            filename,
            content,
            ct,
            timeout=30.0,
        )
        if result.get("ok") and result.get("result", {}).get("photo"):
            photos = result["result"]["photo"]
            largest = max(photos, key=lambda p: p.get("file_size", 0))
            return largest.get("file_id")
    except Exception as e:
        logger.error(f"Ошибка загрузки файла в Telegram: {e}", exc_info=True)
    return None


from api.products.utils.finance_context import get_finance_price_context, FinancePriceContext
from api.products.utils.item_pricing import (
    compute_item_customer_price_rub,
    item_price_rub_base_after_yuan_markup,
)


async def calculate_item_price(item: Item, ctx: FinancePriceContext) -> Decimal:
    """Итоговая цена для покупателя."""
    return compute_item_customer_price_rub(
        item,
        ctx.rate_with_margin,
        ctx.delivery_cost_per_kg,
        yuan_markup_before_rub_percent=ctx.yuan_markup_before_rub_percent,
        customer_price_acquiring_factor=ctx.customer_price_acquiring_factor,
    )


def validate_and_truncate_link(link: Optional[str], item_name: str = "") -> Optional[str]:
    """
    Валидирует и обрезает ссылку до 500 символов
    
    Args:
        link: Ссылка для валидации
        item_name: Название товара (для логирования)
    
    Returns:
        Валидированная и обрезанная ссылка или None
    """
    if not link:
        return None
    
    link_str = str(link).strip()
    
    # Сначала проверяем, что это похоже на URL
    if link_str.startswith("http://") or link_str.startswith("https://"):
        # Это валидный URL - обрезаем до 500 символов если нужно
        if len(link_str) > 500:
            link_str = link_str[:500]
            logger.warning(f"Ссылка обрезана до 500 символов для товара '{item_name}'")
        return link_str
    else:
        # Не URL - проверяем длину
        if len(link_str) > 200:
            # Очень длинная строка и не URL - вероятно ошибка, игнорируем
            logger.warning(f"Подозрительно длинная ссылка (не URL) для товара '{item_name}': {link_str[:100]}...")
            return None
        elif len(link_str) > 500:
            # Длинная строка - обрезаем до 500
            logger.warning(f"Ссылка обрезана до 500 символов для товара '{item_name}'")
            return link_str[:500]
        else:
            # Короткая строка, но не URL - возможно пользователь ошибся, но сохраняем
            return link_str


def normalize_tags(tags: Optional[List[str]]) -> Optional[List[str]]:
    """Нормализует теги: trim, remove empty, unique (с сохранением порядка)."""
    if not tags:
        return None
    normalized: List[str] = []
    seen = set()
    for raw in tags:
        if raw is None:
            continue
        t = str(raw).strip()
        if not t:
            continue
        k = t.lower()
        if k in seen:
            continue
        seen.add(k)
        normalized.append(t)
    return normalized or None


def normalize_optional_name(value: Optional[str], max_len: int = 255) -> Optional[str]:
    """Приводит опциональное текстовое поле к None/обрезанной строке."""
    if value is None:
        return None
    txt = str(value).strip()
    if not txt:
        return None
    return txt[:max_len]


# Pydantic схемы для запросов
class CreateItemRequest(BaseModel):
    name: str
    chinese_name: Optional[str] = None
    description: Optional[str] = None
    price: Decimal
    service_fee_percent: Decimal = Decimal("0")
    estimated_weight_kg: Optional[Decimal] = None
    length_cm: Optional[int] = None
    width_cm: Optional[int] = None
    height_cm: Optional[int] = None
    item_type_id: int
    gender: str
    size: Optional[List[str]] = None
    link: Optional[str] = None
    size_chart_id: Optional[int] = None
    is_legit: Optional[bool] = None
    fixed_price: Optional[Decimal] = None
    tags: Optional[List[str]] = None


class BulkItemData(BaseModel):
    """Данные одного товара для массового создания"""
    name: str
    chinese_name: Optional[str] = None
    description: Optional[str] = None
    link: Optional[str] = None
    photo_indices: Optional[List[int]] = None  # Индексы фото в общем массиве файлов


class BulkCreateItemsRequest(BaseModel):
    """Запрос на массовое создание товаров"""
    # Общие параметры для всех товаров
    price: Decimal
    service_fee_percent: Decimal = Decimal("0")
    estimated_weight_kg: Optional[Decimal] = None
    length_cm: Optional[int] = None
    width_cm: Optional[int] = None
    height_cm: Optional[int] = None
    item_type_id: int
    gender: str
    size: Optional[List[str]] = None
    tags: Optional[List[str]] = None  # Общие теги для всех товаров
    # Массив товаров с индивидуальными данными
    items: List[BulkItemData]


class UpdateItemRequest(BaseModel):
    name: Optional[str] = None
    chinese_name: Optional[str] = None
    description: Optional[str] = None
    price: Optional[Decimal] = None
    service_fee_percent: Optional[Decimal] = None
    estimated_weight_kg: Optional[Decimal] = None
    length_cm: Optional[int] = None
    width_cm: Optional[int] = None
    height_cm: Optional[int] = None
    item_type_id: Optional[int] = None
    gender: Optional[str] = None
    size: Optional[List[str]] = None
    link: Optional[str] = None
    size_chart_id: Optional[int] = None
    is_legit: Optional[bool] = None
    fixed_price: Optional[Decimal] = None
    tags: Optional[List[str]] = None
    @field_validator('description', 'link', 'chinese_name', mode='before')
    @classmethod
    def empty_str_to_none(cls, v):
        """Преобразует пустые строки в None для опциональных полей"""
        if v == "" or (isinstance(v, str) and v.strip() == ""):
            return None
        return v


class ItemDetailResponse(ItemResponse):
    current_price_rub: Optional[Decimal] = None
    service_fee_amount: Optional[Decimal] = None


class ItemPricePreviewRequest(BaseModel):
    """Параметры для расчёта цены без сохранения товара (как в каталоге и карточке)."""

    price: Decimal = Field(..., ge=0, description="Цена в юанях")
    service_fee_percent: Decimal = Field(default=Decimal("0"), ge=0)
    estimated_weight_kg: Optional[Decimal] = Field(default=None, ge=0)


class ItemPricePreviewResponse(BaseModel):
    """Итоговая цена для клиента и наценка (руб), как в списке/карточке админки."""

    current_price_rub: Decimal
    service_fee_amount: Decimal


check_admin_access = require_jwt_permission("catalog")
check_stock_access = require_jwt_permission("catalog")
check_orders_access = require_jwt_permission("orders")


@router.get("/admin/items", response_model=List[ItemResponse])
async def list_items(
    item_type: Optional[str] = Query(None, description="Фильтр по типу товара"),
    gender: Optional[str] = Query(None, description="Фильтр по полу"),
    search: Optional[str] = Query(None, description="Поиск по названию"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Получить список товаров (для админов)"""
    
    query = select(Item).options(joinedload(Item.item_type_rel), joinedload(Item.size_chart))
    
    if item_type:
        # Поддерживаем фильтр по имени типа (для обратной совместимости)
        item_type_result = await session.execute(
            select(ItemType).where(ItemType.name == item_type)
        )
        item_type_obj = item_type_result.scalar_one_or_none()
        if item_type_obj:
            query = query.where(Item.item_type_id == item_type_obj.id)
    if gender:
        query = query.where(Item.gender == gender)
    if search:
        st = search.strip()
        if st:
            conds = [Item.name.ilike(f"%{st}%")]
            if st.isdigit():
                try:
                    sid = int(st)
                    if 0 < sid <= 2147483647:
                        conds.append(Item.id == sid)
                except ValueError:
                    pass
            query = query.where(or_(*conds))
    
    query = query.order_by(Item.id.desc()).offset(skip).limit(limit)
    result = await session.execute(query)
    items = result.unique().scalars().all()
    item_ids_page = [item.id for item in items]
    counts_map = await get_feed_like_dislike_counts_map(session, item_ids_page)

    # Получаем курс валют и стоимость доставки один раз для всех товаров
    ctx = await get_finance_price_context()

    # Загружаем фото для каждого товара и рассчитываем цены
    items_with_photos = []
    for item in items:
        photos_result = await session.execute(
            select(ItemPhoto).where(ItemPhoto.item_id == item.id).order_by(
            func.coalesce(ItemPhoto.sort_order, 999999).asc(),
            ItemPhoto.id
        )
        )
        photos = photos_result.scalars().all()
        
        # Рассчитываем цену в рублях и наценку
        price_rub = await calculate_item_price(item, ctx)
        price_rub_base = item_price_rub_base_after_yuan_markup(
            item, ctx.rate_with_margin, ctx.yuan_markup_before_rub_percent
        )
        service_fee_amount = price_rub_base * (item.service_fee_percent / Decimal("100"))
        
        sc = item.size_chart
        size_chart_resp = SizeChartResponse(id=sc.id, name=sc.name, grid=sc.grid) if sc else None
        lc, dc = like_dislike_for(counts_map, item.id)
        items_with_photos.append(ItemResponse(
            id=item.id,
            name=item.name,
            chinese_name=getattr(item, "chinese_name", None),
            description=item.description,
            price=item.price,
            service_fee_percent=item.service_fee_percent,
            estimated_weight_kg=item.estimated_weight_kg,
            length_cm=getattr(item, "length_cm", None),
            width_cm=getattr(item, "width_cm", None),
            height_cm=getattr(item, "height_cm", None),
            item_type_id=item.item_type_id,
            item_type=item.item_type_rel.name if item.item_type_rel else None,
            gender=item.gender,
            size=item.size if item.size else None,
            link=item.link,
            group_id=item.group_id,
            size_chart_id=getattr(item, "size_chart_id", None),
            size_chart=size_chart_resp,
            photos=[ItemPhotoResponse(
                id=photo.id,
                file_path=photo.file_path,
                telegram_file_id=photo.telegram_file_id,
                vk_attachment=photo.vk_attachment,
                sort_order=getattr(photo, 'sort_order', 0)
            ) for photo in photos],
            price_rub=price_rub,
            service_fee_amount=service_fee_amount,
            is_legit=item.is_legit,
            fixed_price=getattr(item, "fixed_price", None),
            tags=normalize_tags(getattr(item, "tags", None)),
            feed_like_count=lc,
            feed_dislike_count=dc,
        ))
    
    return items_with_photos


class StatsByTypeRow(BaseModel):
    item_type_id: int
    item_type: str
    count: int
    actual_count: int  # без дублей: в группе каждый тип считаем 1 раз
    avg_price_rub: Decimal


@router.get("/admin/items/stats-by-type", response_model=List[StatsByTypeRow])
async def get_items_stats_by_type(
    admin=Depends(check_admin_access),
    session: AsyncSession = Depends(get_session),
):
    """Средняя цена по типам товаров (расчётная цена, фикс не учитывается). actual_count — фактическое кол-во: в группе каждый тип один раз."""
    from collections import defaultdict
    result = await session.execute(
        select(Item).options(joinedload(Item.item_type_rel)).order_by(Item.id)
    )
    items = result.unique().scalars().all()
    ctx = await get_finance_price_context()
    by_type: Dict[int, List[Decimal]] = defaultdict(list)
    actual_by_type: Dict[int, set] = defaultdict(set)  # для каждого типа — множество (group_id или item.id)
    type_names: Dict[int, str] = {}
    for item in items:
        price_rub = await calculate_item_price(item, ctx)
        by_type[item.item_type_id].append(price_rub)
        key = item.group_id if item.group_id is not None else item.id
        actual_by_type[item.item_type_id].add(key)
        if item.item_type_id not in type_names and item.item_type_rel:
            type_names[item.item_type_id] = item.item_type_rel.name
    out = []
    for type_id, prices in by_type.items():
        if not prices:
            continue
        avg = sum(prices, Decimal(0)) / len(prices)
        out.append(StatsByTypeRow(
            item_type_id=type_id,
            item_type=type_names.get(type_id) or str(type_id),
            count=len(prices),
            actual_count=len(actual_by_type[type_id]),
            avg_price_rub=avg,
        ))
    out.sort(key=lambda r: (-r.count, r.item_type))
    return out


@router.post("/admin/items/price-preview", response_model=ItemPricePreviewResponse)
async def preview_item_price(
    request: ItemPricePreviewRequest,
    admin=Depends(check_admin_access),
):
    """
    Расчёт цены для клиента и ориентировочной наценки (сервисный сбор в ₽ от рублёвой базы)
    по текущему курсу и настройкам finance — та же формула, что при отображении товара в админке.
    """
    ctx = await get_finance_price_context()
    stub = Item(
        name="_preview",
        price=request.price,
        service_fee_percent=request.service_fee_percent or Decimal(0),
        estimated_weight_kg=request.estimated_weight_kg,
        item_type_id=1,
        gender="унисекс",
    )
    current_price_rub = await calculate_item_price(stub, ctx)
    price_rub_base = item_price_rub_base_after_yuan_markup(
        stub, ctx.rate_with_margin, ctx.yuan_markup_before_rub_percent
    )
    service_fee_amount = price_rub_base * (stub.service_fee_percent / Decimal("100"))
    return ItemPricePreviewResponse(
        current_price_rub=current_price_rub,
        service_fee_amount=service_fee_amount.quantize(Decimal("0.01")),
    )


@router.get("/admin/items/{item_id}", response_model=ItemDetailResponse)
async def get_item_by_id(
    item_id: int,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Получить детали товара (для админов)"""
    
    result = await session.execute(
        select(Item).options(joinedload(Item.item_type_rel), joinedload(Item.size_chart)).where(Item.id == item_id)
    )
    item = result.unique().scalar_one_or_none()
    
    if not item:
        raise HTTPException(status_code=404, detail="Товар не найден")
    
    # Получаем фото
    photos_result = await session.execute(
        select(ItemPhoto).where(ItemPhoto.item_id == item.id).order_by(
            func.coalesce(ItemPhoto.sort_order, 999999).asc(),
            ItemPhoto.id
        )
    )
    photos = photos_result.scalars().all()
    
    # Рассчитываем текущую цену
    ctx = await get_finance_price_context()
    current_price = await calculate_item_price(item, ctx)
    
    # Рассчитываем сервисный сбор
    price_rub_base = item_price_rub_base_after_yuan_markup(
        item, ctx.rate_with_margin, ctx.yuan_markup_before_rub_percent
    )
    service_fee_amount = price_rub_base * (item.service_fee_percent / Decimal("100"))
    
    sc = item.size_chart
    size_chart_resp = SizeChartResponse(id=sc.id, name=sc.name, grid=sc.grid) if sc else None
    return ItemDetailResponse(
        id=item.id,
        name=item.name,
        chinese_name=getattr(item, "chinese_name", None),
        description=item.description,
        price=item.price,
        service_fee_percent=item.service_fee_percent,
        estimated_weight_kg=item.estimated_weight_kg,
        length_cm=getattr(item, "length_cm", None),
        width_cm=getattr(item, "width_cm", None),
        height_cm=getattr(item, "height_cm", None),
        item_type_id=item.item_type_id,
        item_type=item.item_type_rel.name if item.item_type_rel else None,
        gender=item.gender,
        size=item.size,
        link=item.link,
        size_chart_id=getattr(item, "size_chart_id", None),
        size_chart=size_chart_resp,
        photos=[ItemPhotoResponse(
            id=photo.id,
            file_path=photo.file_path,
            telegram_file_id=photo.telegram_file_id,
            vk_attachment=photo.vk_attachment
        ) for photo in photos],
        current_price_rub=current_price,
        service_fee_amount=service_fee_amount,
        price_rub=current_price,
        is_legit=item.is_legit,
        fixed_price=getattr(item, "fixed_price", None),
        tags=normalize_tags(getattr(item, "tags", None)),
    )


@router.post("/admin/items", response_model=ItemResponse)
async def create_item(
    request: CreateItemRequest,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Создать новый товар (для админов)"""
    
    # Проверяем существование типа вещи
    item_type_result = await session.execute(
        select(ItemType).where(ItemType.id == request.item_type_id)
    )
    item_type = item_type_result.scalar_one_or_none()
    if not item_type:
        raise HTTPException(status_code=400, detail="Тип вещи не найден")
    
    # Валидация и обработка link
    link_value = validate_and_truncate_link(request.link, request.name)
    
    item = Item(
        name=request.name[:255] if len(request.name) > 255 else request.name,  # Ограничиваем длину name
        chinese_name=normalize_optional_name(request.chinese_name),
        description=request.description.strip() if request.description else None,
        price=request.price,
        service_fee_percent=request.service_fee_percent,
        estimated_weight_kg=request.estimated_weight_kg,
        length_cm=request.length_cm,
        width_cm=request.width_cm,
        height_cm=request.height_cm,
        item_type_id=request.item_type_id,
        gender=request.gender,
        size=request.size if request.size else None,
        link=link_value,
        size_chart_id=request.size_chart_id,
        is_legit=request.is_legit if request.is_legit is not None else False,
        fixed_price=request.fixed_price,
        tags=normalize_tags(request.tags),
    )

    session.add(item)
    await session.commit()
    await session.refresh(item)
    await change_item_type_count(session, request.item_type_id, 1)
    await session.commit()
    
    ctx = await get_finance_price_context()
    price_rub = await calculate_item_price(item, ctx)
    from api.products.routers.price_history import upsert_price_history_4h_bucket
    await upsert_price_history_4h_bucket(session, item.id, price_rub)
    await session.commit()
    await invalidate_catalog_cache()
    
    # Если были загружены фото при создании, отправляем их в группу
    # (фото загружаются отдельным запросом после создания товара)
    
    size_chart_resp = None
    if getattr(item, "size_chart_id", None):
        sc_result = await session.execute(select(SizeChart).where(SizeChart.id == item.size_chart_id))
        sc = sc_result.scalar_one_or_none()
        size_chart_resp = SizeChartResponse(id=sc.id, name=sc.name, grid=sc.grid) if sc else None
    return ItemResponse(
        id=item.id,
        name=item.name,
        chinese_name=getattr(item, "chinese_name", None),
        description=item.description,
        price=item.price,
        service_fee_percent=item.service_fee_percent,
        estimated_weight_kg=item.estimated_weight_kg,
        length_cm=item.length_cm,
        width_cm=item.width_cm,
        height_cm=item.height_cm,
        item_type_id=item.item_type_id,
        item_type=item_type.name,
        gender=item.gender,
        size=item.size,
        link=item.link,
        size_chart_id=getattr(item, "size_chart_id", None),
        size_chart=size_chart_resp,
        photos=[],
        is_legit=item.is_legit,
        fixed_price=item.fixed_price,
        tags=normalize_tags(getattr(item, "tags", None)),
    )


@router.post("/admin/items/bulk-create", response_model=List[ItemResponse])
async def bulk_create_items(
    request_data: str = Form(..., description="JSON строка с данными товаров"),
    photos: List[UploadFile] = File(default=[], description="Массив фото для всех товаров (одно и то же имя поля для всех файлов)"),
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """
    Массовое создание товаров
    
    Принимает:
    - request_data: JSON строка с общими параметрами и массивом товаров
    - photos: массив файлов фото (опционально)
    
    Формат request_data:
    {
        "price": 100.00,
        "service_fee_percent": 15.0,
        "estimated_weight_kg": 1.5,
        "item_type_id": 1,
        "gender": "М",
        "size_chart_id": 1,
        "size": ["40", "41", "42"],
        "items": [
            {
                "name": "Товар 1",
                "description": "Описание",
                "link": "https://...",
                "photo_indices": [0, 1]  // индексы фото в массиве photos
            }
        ]
    }
    """
    import json
    
    try:
        # Парсим JSON данные
        data = json.loads(request_data)
    except json.JSONDecodeError as e:
        logger.error(f"Ошибка парсинга JSON: {str(e)}")
        logger.error(f"JSON данные (первые 500 символов): {request_data[:500]}")
        raise HTTPException(status_code=400, detail=f"Неверный формат JSON: {str(e)}")
    
    # Валидируем структуру данных
    if "items" not in data or not isinstance(data["items"], list):
        raise HTTPException(status_code=400, detail="Поле 'items' должно быть массивом")
    
    if not data["items"]:
        raise HTTPException(status_code=400, detail="Массив товаров не может быть пустым")
    
    common_name = (data.get("common_name") or "").strip() if data.get("common_name") else None
    common_link = data.get("common_link")
    common_description = (data.get("common_description") or "").strip() if data.get("common_description") else None
    use_common_name = bool(common_name)

    # Валидация каждого товара
    for idx, item_data in enumerate(data["items"]):
        if not isinstance(item_data, dict):
            raise HTTPException(status_code=400, detail=f"Товар #{idx + 1} должен быть объектом")
        if not use_common_name and (not item_data.get("name") or not str(item_data.get("name", "")).strip()):
            raise HTTPException(status_code=400, detail=f"Товар #{idx + 1}: название обязательно (или включите режим «один товар на все» с общим названием)")
        # Валидация link
        if not use_common_name and "link" in item_data and item_data["link"]:
            link_str = str(item_data["link"])
            if len(link_str) > 500:
                logger.warning(f"Товар #{idx + 1}: ссылка слишком длинная ({len(link_str)} символов), обрезаем до 500")
                item_data["link"] = link_str[:500]
    
    # Проверяем существование типа вещи
    item_type_id = data.get("item_type_id")
    if not item_type_id:
        raise HTTPException(status_code=400, detail="item_type_id обязателен")
    
    item_type_result = await session.execute(
        select(ItemType).where(ItemType.id == item_type_id)
    )
    item_type = item_type_result.scalar_one_or_none()
    if not item_type:
        raise HTTPException(status_code=400, detail="Тип вещи не найден")
    
    # Получаем общие параметры
    common_price = Decimal(str(data.get("price", 0)))
    common_service_fee_percent = Decimal(str(data.get("service_fee_percent", 0)))
    common_estimated_weight_kg = Decimal(str(data["estimated_weight_kg"])) if data.get("estimated_weight_kg") else None
    common_length_cm = data.get("length_cm")
    common_width_cm = data.get("width_cm")
    common_height_cm = data.get("height_cm")
    common_gender = data.get("gender")
    common_size = data.get("size")
    common_size_chart_id = data.get("size_chart_id")
    common_tags = normalize_tags(data.get("tags"))  # Общие теги для всех товаров
    if common_size_chart_id is not None:
        sc_result = await session.execute(select(SizeChart).where(SizeChart.id == common_size_chart_id))
        if not sc_result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Размерная сетка не найдена")
    
    if not common_gender:
        raise HTTPException(status_code=400, detail="gender обязателен")
    
    ctx = await get_finance_price_context()
    from api.products.routers.price_history import upsert_price_history_4h_bucket

    # Опционально создаём группу (все созданные товары попадут в неё)
    group_id = None
    group_name = (data.get("group_name") or "").strip()
    if group_name:
        group = ItemGroup(name=group_name[:255] if len(group_name) > 255 else group_name)
        session.add(group)
        await session.flush()
        group_id = group.id

    created_items = []
    
    for item_data in data["items"]:
        if use_common_name:
            item_name = common_name[:255] if len(common_name) > 255 else common_name
            link_value = validate_and_truncate_link(common_link, item_name)
            item_description = common_description or None
        else:
            item_name = item_data.get("name", "").strip()
            item_name = item_name[:255] if len(item_name) > 255 else item_name
            link_value = validate_and_truncate_link(item_data.get("link"), item_name)
            item_description = item_data.get("description", "").strip() if item_data.get("description") else None
        item_chinese_name = normalize_optional_name(item_data.get("chinese_name"))
        
        # Создаем товар
        item = Item(
            name=item_name,
            chinese_name=item_chinese_name,
            description=item_description,
            price=common_price,
            service_fee_percent=common_service_fee_percent,
            estimated_weight_kg=common_estimated_weight_kg,
            length_cm=common_length_cm,
            width_cm=common_width_cm,
            height_cm=common_height_cm,
            item_type_id=item_type_id,
            gender=common_gender,
            size=common_size if common_size else None,
            link=link_value,
            size_chart_id=common_size_chart_id if common_size_chart_id else None,
            is_legit=False,
            tags=common_tags,
            group_id=group_id,
        )
        
        session.add(item)
        await session.commit()
        await session.refresh(item)
        price_rub = await calculate_item_price(item, ctx)
        await upsert_price_history_4h_bucket(session, item.id, price_rub)
        await session.commit()
        
        # Загружаем фото для товара
        photo_indices = item_data.get("photo_indices", [])
        if photo_indices and photos:
            item_photos = []
            for photo_idx in photo_indices:
                if 0 <= photo_idx < len(photos):
                    photo_file = photos[photo_idx]

                    # Сохраняем файл сжатым
                    file_path = await save_compressed_image(photo_file)
                    file_name = file_path.name
                    
                    # Определяем sort_order
                    max_sort_order_result = await session.execute(
                        select(func.coalesce(func.max(ItemPhoto.sort_order), -1)).where(ItemPhoto.item_id == item.id)
                    )
                    max_sort_order = max_sort_order_result.scalar_one_or_none() or -1
                    new_sort_order = max_sort_order + 1
                    
                    # Создаем запись в БД
                    relative_path = f"uploads/items/{file_name}"
                    item_photo = ItemPhoto(
                        item_id=item.id,
                        file_path=relative_path,
                        telegram_file_id=None,
                        vk_attachment=None,
                        sort_order=new_sort_order
                    )
                    
                    session.add(item_photo)
                    await session.commit()
                    await session.refresh(item_photo)
                    
                    # Отправляем фото в Telegram группу для получения file_id
                    try:
                        file_id = await send_photo_to_telegram_group_and_get_file_id(item_photo.id, str(file_path))
                        if file_id:
                            item_photo.telegram_file_id = file_id
                            await session.commit()
                            await session.refresh(item_photo)
                    except Exception as e:
                        logger.error(f"Ошибка при отправке фото ID={item_photo.id} в Telegram: {e}", exc_info=True)
                    
                    item_photos.append(item_photo)
            
            # Загружаем фото для ответа
            photos_result = await session.execute(
                select(ItemPhoto).where(ItemPhoto.item_id == item.id).order_by(
                    func.coalesce(ItemPhoto.sort_order, 999999).asc(),
                    ItemPhoto.id
                )
            )
            item_photos_list = photos_result.scalars().all()
        else:
            item_photos_list = []
        
        # Формируем ответ
        created_items.append(ItemResponse(
            id=item.id,
            name=item.name,
            chinese_name=getattr(item, "chinese_name", None),
            description=item.description,
            price=item.price,
            service_fee_percent=item.service_fee_percent,
            estimated_weight_kg=item.estimated_weight_kg,
            length_cm=item.length_cm,
            width_cm=item.width_cm,
            height_cm=item.height_cm,
            item_type_id=item.item_type_id,
            item_type=item_type.name,
            gender=item.gender,
            size=item.size,
            link=item.link,
            photos=[
                ItemPhotoResponse(
                    id=photo.id,
                    file_path=photo.file_path,
                    telegram_file_id=photo.telegram_file_id,
                    vk_attachment=photo.vk_attachment,
                    sort_order=photo.sort_order
                )
                for photo in item_photos_list
            ],
            is_legit=item.is_legit,
            tags=normalize_tags(getattr(item, "tags", None)),
        ))
    
    await change_item_type_count(session, item_type_id, len(created_items))
    await session.commit()
    await invalidate_catalog_cache()
    return created_items


@router.put("/admin/items/{item_id}", response_model=ItemResponse)
async def update_item(
    item_id: int,
    request: UpdateItemRequest,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Обновить товар (для админов)"""
    
    result = await session.execute(select(Item).where(Item.id == item_id))
    item = result.scalar_one_or_none()
    
    if not item:
        raise HTTPException(status_code=404, detail="Товар не найден")
    
    # Отслеживаем, изменились ли параметры, влияющие на цену (для обновления истории)
    price_affecting_fields_changed = False
    
    # Обновляем поля
    # Используем model_fields_set (Pydantic v2) или __fields_set__ (Pydantic v1) чтобы определить, какие поля были переданы
    fields_set = getattr(request, 'model_fields_set', None) or getattr(request, '__fields_set__', None) or set()
    
    if request.name is not None:
        item.name = request.name
    if 'chinese_name' in fields_set:
        item.chinese_name = normalize_optional_name(request.chinese_name)
    if 'description' in fields_set:
        # Поле было передано - валидатор уже преобразовал пустые строки в None
        item.description = request.description
    if request.price is not None:
        if item.price != request.price:
            price_affecting_fields_changed = True
        item.price = request.price
    if request.service_fee_percent is not None:
        if item.service_fee_percent != request.service_fee_percent:
            price_affecting_fields_changed = True
        item.service_fee_percent = request.service_fee_percent
    if request.estimated_weight_kg is not None:
        if item.estimated_weight_kg != request.estimated_weight_kg:
            price_affecting_fields_changed = True
        item.estimated_weight_kg = request.estimated_weight_kg
    if request.item_type_id is not None:
        # Проверяем существование типа вещи
        item_type_result = await session.execute(
            select(ItemType).where(ItemType.id == request.item_type_id)
        )
        item_type = item_type_result.scalar_one_or_none()
        if not item_type:
            raise HTTPException(status_code=400, detail="Тип вещи не найден")
        old_type_id = item.item_type_id
        item.item_type_id = request.item_type_id
        if old_type_id != request.item_type_id:
            await change_item_type_count(session, old_type_id, -1)
            await change_item_type_count(session, request.item_type_id, 1)
    if request.gender is not None:
        item.gender = request.gender
    if 'size' in fields_set:
        # Поле было передано - валидатор уже преобразовал пустые строки в None
        item.size = request.size
    if 'link' in fields_set:
        # Поле было передано - валидатор уже преобразовал пустые строки в None
        # Валидация и обработка link
        item.link = validate_and_truncate_link(request.link, item.name or f"ID {item_id}")
    if 'is_legit' in fields_set:
        item.is_legit = request.is_legit
    if 'fixed_price' in fields_set:
        item.fixed_price = request.fixed_price
    if request.length_cm is not None:
        item.length_cm = request.length_cm
    if request.width_cm is not None:
        item.width_cm = request.width_cm
    if request.height_cm is not None:
        item.height_cm = request.height_cm
    if "size_chart_id" in fields_set:
        item.size_chart_id = request.size_chart_id
    if "tags" in fields_set:
        item.tags = normalize_tags(request.tags)
    
    await session.commit()
    await session.refresh(item)
    await invalidate_catalog_cache()
    
    # Если изменились параметры, влияющие на итоговую цену - обновляем историю
    if price_affecting_fields_changed:
        ctx = await get_finance_price_context()
        new_price_rub = await calculate_item_price(item, ctx)
        from api.products.routers.price_history import upsert_price_history_4h_bucket
        await upsert_price_history_4h_bucket(session, item.id, new_price_rub)
        await session.commit()
    
    # Получаем фото и тип вещи
    photos_result = await session.execute(
        select(ItemPhoto).where(ItemPhoto.item_id == item.id).order_by(
            func.coalesce(ItemPhoto.sort_order, 999999).asc(),
            ItemPhoto.id
        )
    )
    photos = photos_result.scalars().all()
    
    # Загружаем тип вещи
    item_type_result = await session.execute(
        select(ItemType).where(ItemType.id == item.item_type_id)
    )
    item_type = item_type_result.scalar_one_or_none()
    
    size_chart_resp = None
    if getattr(item, "size_chart_id", None):
        sc_result = await session.execute(select(SizeChart).where(SizeChart.id == item.size_chart_id))
        sc = sc_result.scalar_one_or_none()
        size_chart_resp = SizeChartResponse(id=sc.id, name=sc.name, grid=sc.grid) if sc else None
    return ItemResponse(
        id=item.id,
        name=item.name,
        chinese_name=getattr(item, "chinese_name", None),
        description=item.description,
        price=item.price,
        service_fee_percent=item.service_fee_percent,
        estimated_weight_kg=item.estimated_weight_kg,
        length_cm=getattr(item, "length_cm", None),
        width_cm=getattr(item, "width_cm", None),
        height_cm=getattr(item, "height_cm", None),
        item_type_id=item.item_type_id,
        item_type=item_type.name if item_type else None,
        gender=item.gender,
        size=item.size,
        link=item.link,
        size_chart_id=getattr(item, "size_chart_id", None),
        size_chart=size_chart_resp,
        photos=[ItemPhotoResponse(
            id=photo.id,
            file_path=photo.file_path,
            telegram_file_id=photo.telegram_file_id,
            vk_attachment=photo.vk_attachment,
            sort_order=photo.sort_order,
        ) for photo in photos],
        fixed_price=getattr(item, "fixed_price", None),
        tags=normalize_tags(getattr(item, "tags", None)),
    )


class BulkUpdateItemsRequest(BaseModel):
    """Массовое обновление товаров: перезапись выбранных полей у выбранных товаров."""
    item_ids: List[int]
    is_legit: Optional[bool] = None
    size_chart_id: Optional[int] = None  # None — сбросить сетку
    item_type_id: Optional[int] = None
    gender: Optional[str] = None  # М, Ж, унисекс
    size: Optional[List[str]] = None  # массив размеров (перезапись)
    price: Optional[Decimal] = None  # цена в юанях
    service_fee_percent: Optional[Decimal] = None  # % наценки
    link: Optional[str] = None  # ссылка на товар (перезапись)
    tags: Optional[List[str]] = None  # теги для поиска (перезапись)
    add_tags: Optional[List[str]] = None  # теги добавить к существующим (объединение без дубликатов)


VALID_GENDERS = {"М", "Ж", "унисекс"}


@router.patch("/admin/items/bulk-update", response_model=dict)
async def bulk_update_items(
    request: BulkUpdateItemsRequest,
    admin=Depends(check_admin_access),
    session: AsyncSession = Depends(get_session),
):
    """Массово перезаписать выбранные поля у выбранных товаров."""
    if not request.item_ids:
        raise HTTPException(status_code=400, detail="Укажите хотя бы один товар (item_ids)")
    provided = request.model_dump(exclude_unset=True)
    allowed = {"is_legit", "size_chart_id", "item_type_id", "gender", "size", "price", "service_fee_percent", "link", "tags", "add_tags"}
    payload = {k: v for k, v in provided.items() if k in allowed and k != "item_ids"}
    if not payload:
        raise HTTPException(status_code=400, detail="Укажите хотя бы одно поле для обновления")
    if "gender" in payload and payload["gender"] not in VALID_GENDERS:
        raise HTTPException(status_code=400, detail="gender должен быть: М, Ж или унисекс")
    if "item_type_id" in payload and payload["item_type_id"] is not None:
        type_result = await session.execute(select(ItemType).where(ItemType.id == payload["item_type_id"]))
        if not type_result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Тип товара не найден")
    if "size_chart_id" in payload and payload["size_chart_id"] is not None:
        sc_result = await session.execute(select(SizeChart).where(SizeChart.id == payload["size_chart_id"]))
        if not sc_result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Размерная сетка не найдена")
    result = await session.execute(select(Item).where(Item.id.in_(request.item_ids)))
    items = result.scalars().all()
    price_updated = "price" in payload or "service_fee_percent" in payload
    ctx = None
    if price_updated:
        ctx = await get_finance_price_context()
    from api.products.routers.price_history import upsert_price_history_4h_bucket

    for item in items:
        if "is_legit" in payload and payload["is_legit"] is not None:
            item.is_legit = payload["is_legit"]
        if "size_chart_id" in payload:
            item.size_chart_id = payload["size_chart_id"] if payload["size_chart_id"] else None
        if "item_type_id" in payload and payload["item_type_id"] is not None:
            new_tid = payload["item_type_id"]
            old_tid = item.item_type_id
            item.item_type_id = new_tid
            if old_tid != new_tid:
                await change_item_type_count(session, old_tid, -1)
                await change_item_type_count(session, new_tid, 1)
        if "gender" in payload and payload["gender"] is not None:
            item.gender = payload["gender"]
        if "size" in payload:
            item.size = payload["size"] if payload["size"] else None
        if "price" in payload and payload["price"] is not None:
            item.price = payload["price"]
        if "service_fee_percent" in payload and payload["service_fee_percent"] is not None:
            item.service_fee_percent = payload["service_fee_percent"]
        if "link" in payload:
            item.link = validate_and_truncate_link(payload.get("link"), item.name or f"ID {item.id}")
        if "tags" in payload:
            item.tags = payload["tags"] if payload.get("tags") else None
        if "add_tags" in payload and payload["add_tags"]:
            existing = list(item.tags) if item.tags else []
            merged = list(dict.fromkeys(existing + [t.strip() for t in payload["add_tags"] if t and str(t).strip()]))
            item.tags = merged if merged else None
    await session.commit()
    if price_updated and ctx is not None:
        for item in items:
            await session.refresh(item)
            price_rub = await calculate_item_price(item, ctx)
            await upsert_price_history_4h_bucket(session, item.id, price_rub)
        await session.commit()
    await invalidate_catalog_cache()
    return {"updated_count": len(items), "item_ids": request.item_ids}


@router.delete("/admin/items/{item_id}")
async def delete_item(
    item_id: int,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Удалить товар (для админов)"""
    
    result = await session.execute(select(Item).where(Item.id == item_id))
    item = result.scalar_one_or_none()
    
    if not item:
        raise HTTPException(status_code=404, detail="Товар не найден")
    
    type_id = item.item_type_id
    await session.delete(item)
    await session.commit()
    await change_item_type_count(session, type_id, -1)
    await session.commit()
    await invalidate_catalog_cache()
    return {"message": "Товар удален", "item_id": item_id}


# --- Склад: остатки и поставки ---

class StockRowResponse(BaseModel):
    """Одна строка остатков: товар, размер, количество и резерв"""
    item_id: int
    item_name: str
    size: str
    quantity: int
    reserved_quantity: int = 0


class StockAdjustRequest(BaseModel):
    """Списание/корректировка: размер и дельта (отрицательная — списание)"""
    size: str
    delta: int  # может быть отрицательным


class SupplyItemRequest(BaseModel):
    """Одна позиция поставки"""
    item_id: int
    size: str
    quantity: int


class SupplyRequest(BaseModel):
    """Новая поставка: список позиций"""
    items: List[SupplyItemRequest]


@router.get("/admin/stock", response_model=List[StockRowResponse])
async def list_stock(
    item_id: Optional[int] = Query(None, description="Фильтр по товару"),
    search: Optional[str] = Query(None, description="Поиск по части названия товара"),
    admin=Depends(check_stock_access),
    session: AsyncSession = Depends(get_session),
):
    """Список остатков на складе.

    Показывает:
    - общее количество по размеру (quantity из ItemStock);
    - количество в активном резерве (reserved_quantity из ItemReservation со статусом 'active').

    search — поиск по части названия (без учёта регистра).
    """
    # Подзапрос по резервам: сумма активных резервов по (item_id, size)
    reservations_subq = (
        select(
            ItemReservation.item_id.label("item_id"),
            ItemReservation.size.label("size"),
            func.coalesce(func.sum(ItemReservation.quantity), 0).label("reserved_quantity"),
        )
        .where(ItemReservation.status == "active")
        .group_by(ItemReservation.item_id, ItemReservation.size)
        .subquery()
    )

    query = (
        select(
            ItemStock,
            Item.name,
            func.coalesce(reservations_subq.c.reserved_quantity, 0).label("reserved_quantity"),
        )
        .join(Item, Item.id == ItemStock.item_id)
        .outerjoin(
            reservations_subq,
            and_(
                reservations_subq.c.item_id == ItemStock.item_id,
                reservations_subq.c.size == ItemStock.size,
            ),
        )
        .order_by(Item.id, ItemStock.size)
    )
    if item_id is not None:
        query = query.where(ItemStock.item_id == item_id)
    if search and search.strip():
        query = query.where(Item.name.ilike(f"%{search.strip()}%"))

    result = await session.execute(query)
    rows = result.all()
    return [
        StockRowResponse(
            item_id=s.item_id,
            item_name=name,
            size=s.size,
            quantity=s.quantity,
            reserved_quantity=int(reserved_qty or 0),
        )
        for s, name, reserved_qty in rows
    ]


@router.post("/admin/stock/{item_id}/adjust", response_model=StockRowResponse)
async def adjust_stock(
    item_id: int,
    body: StockAdjustRequest,
    admin=Depends(check_stock_access),
    session: AsyncSession = Depends(get_session),
):
    """Ручная корректировка остатка (списание или приход). Создаёт запись по размеру, если её не было."""
    item_result = await session.execute(select(Item).where(Item.id == item_id))
    item = item_result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Товар не найден")
    stock_result = await session.execute(
        select(ItemStock).where(and_(ItemStock.item_id == item_id, ItemStock.size == body.size))
    )
    row = stock_result.scalar_one_or_none()
    if row is None:
        if body.delta < 0:
            raise HTTPException(status_code=400, detail="Нельзя списать больше, чем есть (остаток по размеру 0)")
        row = ItemStock(item_id=item_id, size=body.size, quantity=body.delta)
        session.add(row)
    else:
        row.quantity += body.delta
        if row.quantity < 0:
            raise HTTPException(status_code=400, detail="Нельзя списать больше, чем есть")
    await session.commit()
    await session.refresh(row)
    return StockRowResponse(item_id=row.item_id, item_name=item.name, size=row.size, quantity=row.quantity)


@router.post("/admin/stock/supply", response_model=List[StockRowResponse])
async def supply_stock(
    body: SupplyRequest,
    admin=Depends(check_stock_access),
    session: AsyncSession = Depends(get_session),
):
    """Новая поставка: добавить количество по выбранным товарам/размерам."""
    if not body.items:
        return []
    out = []
    for req in body.items:
        item_result = await session.execute(select(Item).where(Item.id == req.item_id))
        item = item_result.scalar_one_or_none()
        if not item:
            raise HTTPException(status_code=404, detail=f"Товар с id={req.item_id} не найден")
        stock_result = await session.execute(
            select(ItemStock).where(and_(ItemStock.item_id == req.item_id, ItemStock.size == req.size))
        )
        row = stock_result.scalar_one_or_none()
        if row is None:
            if req.quantity <= 0:
                raise HTTPException(status_code=400, detail="Количество в поставке должно быть > 0")
            row = ItemStock(item_id=req.item_id, size=req.size, quantity=req.quantity)
            session.add(row)
        else:
            row.quantity += req.quantity
            if row.quantity < 0:
                raise HTTPException(status_code=400, detail="Итоговый остаток не может быть отрицательным")
        await session.flush()
        out.append(StockRowResponse(item_id=row.item_id, item_name=item.name, size=row.size, quantity=row.quantity))
    await session.commit()
    return out


class ReservationResponse(BaseModel):
    id: int
    item_id: int
    item_name: str
    size: str
    quantity: int
    user_id: int
    created_at: Optional[str] = None
    status: str


@router.get("/admin/reservations", response_model=List[ReservationResponse])
async def list_reservations(
    status_filter: Optional[str] = Query("active", description="active, cancelled, used; по умолчанию active — только активные (завершённые не показываются)"),
    item_id: Optional[int] = Query(None),
    admin=Depends(check_stock_access),
    session: AsyncSession = Depends(get_session),
):
    """Список резервирований. По умолчанию только активные (used после завершения заказа не отображаются)."""
    query = select(ItemReservation, Item.name).join(Item, Item.id == ItemReservation.item_id).order_by(ItemReservation.created_at.desc())
    if status_filter:
        query = query.where(ItemReservation.status == status_filter)
    if item_id is not None:
        query = query.where(ItemReservation.item_id == item_id)
    result = await session.execute(query)
    rows = result.all()
    return [
        ReservationResponse(
            id=r.id,
            item_id=r.item_id,
            item_name=name,
            size=r.size,
            quantity=r.quantity,
            user_id=r.user_id,
            created_at=r.created_at.isoformat() if getattr(r.created_at, "isoformat", None) else None,
            status=r.status,
        )
        for r, name in rows
    ]


class ReservationStatusUpdate(BaseModel):
    status: str  # "cancelled" | "used"


@router.patch("/admin/reservations/{reservation_id}")
async def update_reservation_status(
    reservation_id: int,
    body: ReservationStatusUpdate,
    admin=Depends(check_stock_access),
    session: AsyncSession = Depends(get_session),
):
    """Отменить или отметить резерв как использованный."""
    if body.status not in ("cancelled", "used"):
        raise HTTPException(status_code=400, detail="status должен быть cancelled или used")
    result = await session.execute(select(ItemReservation).where(ItemReservation.id == reservation_id))
    r = result.scalar_one_or_none()
    if not r:
        raise HTTPException(status_code=404, detail="Резервирование не найдено")
    r.status = body.status
    await session.commit()
    return {"success": True, "reservation_id": reservation_id, "status": body.status}


@router.post("/admin/items/{item_id}/photos", response_model=ItemPhotoResponse)
async def add_item_photo(
    item_id: int,
    photo: UploadFile = File(...),
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Добавить фотографию к товару (для админов)"""
    
    # Проверяем существование товара
    result = await session.execute(select(Item).where(Item.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Товар не найден")
    
    # Проверяем количество фото (максимум 10)
    photos_count_result = await session.execute(
        select(ItemPhoto).where(ItemPhoto.item_id == item_id)
    )
    existing_photos = photos_count_result.scalars().all()
    if len(existing_photos) >= 10:
        raise HTTPException(status_code=400, detail="Максимум 10 фотографий на товар")
    
    # Определяем sort_order (максимальный + 1 или 0, если фото нет)
    # Используем COALESCE для обработки NULL значений
    max_sort_order_result = await session.execute(
        select(func.coalesce(func.max(ItemPhoto.sort_order), -1)).where(ItemPhoto.item_id == item_id)
    )
    max_sort_order = max_sort_order_result.scalar_one_or_none() or -1
    new_sort_order = max_sort_order + 1
    
    # Сохраняем файл сжатым
    file_path = await save_compressed_image(photo)
    file_name = file_path.name
    
    # Создаем запись в БД
    relative_path = f"uploads/items/{file_name}"
    item_photo = ItemPhoto(
        item_id=item_id,
        file_path=relative_path,
        telegram_file_id=None,  # Будет заполнено позже через механизм отправки в чат
        vk_attachment=None,
        sort_order=new_sort_order
    )
    
    session.add(item_photo)
    await session.commit()
    await session.refresh(item_photo)
    
    logger.info(f"Создано фото ID={item_photo.id}, путь={relative_path}, абсолютный путь={file_path}")
    
    # Отправляем фото в Telegram группу для получения file_id (VK attachment загружается при первой отправке в боте)
    try:
        # Telegram file_id
        file_id = await send_photo_to_telegram_group_and_get_file_id(item_photo.id, str(file_path))
        if file_id:
            item_photo.telegram_file_id = file_id
            await session.commit()
            await session.refresh(item_photo)
            logger.info(f"✅ File ID {file_id[:30]}... сохранен для фото ID={item_photo.id}")
        else:
            logger.warning(f"⚠️ File ID не получен для фото ID={item_photo.id}")
    except Exception as e:
        logger.error(f"Ошибка при отправке фото ID={item_photo.id} в Telegram группу: {e}", exc_info=True)
    
    await session.refresh(item_photo)
    
    return ItemPhotoResponse(
        id=item_photo.id,
        file_path=item_photo.file_path,
        telegram_file_id=item_photo.telegram_file_id,
        vk_attachment=item_photo.vk_attachment,
        sort_order=item_photo.sort_order
    )


class PhotoUpdateRequest(BaseModel):
    """Схема для обновления фото"""
    vk_attachment: Optional[str] = None
    sort_order: Optional[int] = None


def verify_internal_token(x_internal_token: Optional[str] = Header(None, alias="X-Internal-Token")):
    """Проверка внутреннего токена для межсервисного взаимодействия"""
    if not x_internal_token:
        raise HTTPException(status_code=401, detail="Требуется внутренний токен")
    clean_token = x_internal_token.replace("Bearer ", "").strip() if x_internal_token.startswith("Bearer") else x_internal_token.strip()
    if clean_token != INTERNAL_TOKEN:
        raise HTTPException(status_code=401, detail="Невалидный внутренний токен")
    return True


@router.patch("/internal/items/photos/{photo_id}/vk-attachment")
async def update_photo_vk_attachment_internal(
    photo_id: int,
    update_data: PhotoUpdateRequest,
    _ = Depends(verify_internal_token),
    session: AsyncSession = Depends(get_session)
):
    """Обновить vk_attachment для фото (внутренний endpoint для ботов)"""
    
    result = await session.execute(select(ItemPhoto).where(ItemPhoto.id == photo_id))
    photo = result.scalar_one_or_none()
    
    if not photo:
        raise HTTPException(status_code=404, detail="Фотография не найдена")
    
    # Обновляем vk_attachment если передан
    if update_data.vk_attachment is not None:
        photo.vk_attachment = update_data.vk_attachment
        await session.commit()
        await session.refresh(photo)
    
    return ItemPhotoResponse(
        id=photo.id,
        file_path=photo.file_path,
        telegram_file_id=photo.telegram_file_id,
        vk_attachment=photo.vk_attachment,
        sort_order=getattr(photo, 'sort_order', 0)
    )


@router.patch("/admin/items/photos/{photo_id}")
async def update_item_photo(
    photo_id: int,
    update_data: PhotoUpdateRequest,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Обновить фотографию товара (для админов) - обновить vk_attachment или sort_order"""
    
    result = await session.execute(select(ItemPhoto).where(ItemPhoto.id == photo_id))
    photo = result.scalar_one_or_none()
    
    if not photo:
        raise HTTPException(status_code=404, detail="Фотография не найдена")
    
    # Обновляем vk_attachment если передан
    if update_data.vk_attachment is not None:
        photo.vk_attachment = update_data.vk_attachment
    # Обновляем sort_order если передан
    if update_data.sort_order is not None:
        photo.sort_order = update_data.sort_order
    
    if update_data.vk_attachment is not None or update_data.sort_order is not None:
        await session.commit()
        await session.refresh(photo)
    
    return ItemPhotoResponse(
        id=photo.id,
        file_path=photo.file_path,
        telegram_file_id=photo.telegram_file_id,
        vk_attachment=photo.vk_attachment,
        sort_order=getattr(photo, 'sort_order', 0)
    )


@router.delete("/admin/items/photos/{photo_id}")
async def delete_item_photo(
    photo_id: int,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Удалить фотографию товара (для админов)"""
    
    result = await session.execute(select(ItemPhoto).where(ItemPhoto.id == photo_id))
    photo = result.scalar_one_or_none()
    
    if not photo:
        raise HTTPException(status_code=404, detail="Фотография не найдена")
    
    item_id = photo.item_id
    
    # Удаляем файл (если существует)
    file_path = Path(photo.file_path)
    if file_path.exists():
        try:
            file_path.unlink()
        except Exception:
            pass  # Игнорируем ошибки удаления файла
    
    await session.delete(photo)
    await session.commit()
    
    # Перенумеровываем sort_order оставшихся фотографий
    remaining_photos_result = await session.execute(
        select(ItemPhoto).where(ItemPhoto.item_id == item_id).order_by(
            func.coalesce(ItemPhoto.sort_order, 999999).asc(),
            ItemPhoto.id
        )
    )
    remaining_photos = remaining_photos_result.scalars().all()
    
    for index, remaining_photo in enumerate(remaining_photos):
        if remaining_photo.sort_order != index:
            remaining_photo.sort_order = index
    
    await session.commit()
    
    return {"message": "Фотография удалена", "photo_id": photo_id}


class ReorderPhotosRequest(BaseModel):
    """Запрос на изменение порядка фотографий"""
    photo_ids: List[int]  # Список ID фотографий в нужном порядке


@router.post("/admin/items/{item_id}/photos/reorder")
async def reorder_item_photos(
    item_id: int,
    request: ReorderPhotosRequest,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Изменить порядок фотографий товара"""
    
    # Проверяем существование товара
    result = await session.execute(select(Item).where(Item.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Товар не найден")
    
    # Проверяем, что все фото принадлежат этому товару
    photos_result = await session.execute(
        select(ItemPhoto).where(
            ItemPhoto.item_id == item_id,
            ItemPhoto.id.in_(request.photo_ids)
        )
    )
    photos = photos_result.scalars().all()
    
    if len(photos) != len(request.photo_ids):
        raise HTTPException(status_code=400, detail="Некоторые фотографии не найдены или не принадлежат этому товару")
    
    # Создаем маппинг photo_id -> index (порядок)
    photo_map = {photo.id: photo for photo in photos}
    
    # Обновляем sort_order для каждой фотографии
    for index, photo_id in enumerate(request.photo_ids):
        if photo_id in photo_map:
            photo_map[photo_id].sort_order = index
    
    await session.commit()
    
    # Возвращаем обновленные фотографии
    photos_result = await session.execute(
        select(ItemPhoto).where(ItemPhoto.item_id == item_id).order_by(
            func.coalesce(ItemPhoto.sort_order, 999999).asc(),
            ItemPhoto.id
        )
    )
    updated_photos = photos_result.scalars().all()
    
    return {
        "message": "Порядок фотографий обновлен",
        "photos": [ItemPhotoResponse(
            id=photo.id,
            file_path=photo.file_path,
            telegram_file_id=photo.telegram_file_id,
            vk_attachment=photo.vk_attachment,
            sort_order=getattr(photo, 'sort_order', 0)
        ) for photo in updated_photos]
    }


@router.post("/admin/items/photos/update-ids")
async def update_photo_ids(
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Обновить telegram_file_id для всех фотографий, у которых его нет. VK attachment загружается автоматически при первой отправке в VK боте."""
    
    # Получаем все фото, у которых нет telegram_file_id
    # VK attachment загружается автоматически при первой отправке в VK боте
    result = await session.execute(
        select(ItemPhoto).where(ItemPhoto.telegram_file_id.is_(None))
    )
    photos_to_update = result.scalars().all()
    
    if not photos_to_update:
        return {
            "message": "Все фотографии уже имеют file_id",
            "updated_count": 0,
            "telegram_updated": 0
        }
    
    telegram_updated = 0
    
    for photo in photos_to_update:
        file_path = Path(photo.file_path)
        
        # Проверяем существование файла
        if not file_path.exists():
            # Пробуем разные варианты путей
            possible_paths = [
                Path(f"/app/{photo.file_path}"),
                Path(photo.file_path),
                Path(f"uploads/items/{file_path.name}"),
                Path(f"/app/uploads/items/{file_path.name}")
            ]
            for path in possible_paths:
                if path.exists():
                    file_path = path
                    break
            else:
                logger.warning(f"Файл не найден для фото ID={photo.id}: {photo.file_path}")
                continue
        
        # Обновляем Telegram file_id, если его нет
        if not photo.telegram_file_id:
            try:
                file_id = await send_photo_to_telegram_group_and_get_file_id(photo.id, str(file_path))
                if file_id:
                    photo.telegram_file_id = file_id
                    telegram_updated += 1
                    logger.info(f"✅ Обновлен Telegram file_id для фото ID={photo.id}")
            except Exception as e:
                logger.error(f"Ошибка при обновлении Telegram file_id для фото ID={photo.id}: {e}", exc_info=True)
        
        # VK attachment загружается автоматически при первой отправке фото в VK боте
        
        await session.commit()
        await session.refresh(photo)
    
    return {
        "message": "Обновление завершено",
        "updated_count": len(photos_to_update),
        "telegram_updated": telegram_updated
    }


# ========== Эндпоинт для отправки поста в Telegram канал ==========

TG_CAPTION_MAX_LENGTH = 1024


def get_telegram_caption_length(html: str) -> int:
    """Длина caption как считает Telegram после парсинга HTML. <a href="...">text</a> -> только text."""
    if not html:
        return 0
    # Заменяем <a href="...">текст</a> на текст
    s = re.sub(r'<a\s+href="[^"]*">([^<]*)</a>', r'\1', html, flags=re.IGNORECASE)
    # Удаляем оставшиеся теги
    s = re.sub(r'<[^>]+>', '', s)
    return len(s)


def _parse_photo_ids(photo_ids_str: str | None) -> list[int] | None:
    """Парсит photo_ids из JSON строки."""
    if not photo_ids_str or photo_ids_str.strip() in ("", "null"):
        return None
    try:
        data = __import__("json").loads(photo_ids_str)
        return data if isinstance(data, list) else None
    except Exception:
        return None


@router.post("/admin/items/{item_id}/post-to-telegram")
async def post_item_to_telegram(
    item_id: int,
    message_text: str = Form(...),
    photo_ids: str = Form("null"),
    additional_photos: list[UploadFile] = File(default=[]),
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    raise HTTPException(
        status_code=410,
        detail="Публикация в Telegram-канал отключена в StoreKPLite.",
    )


# ========== Админские эндпоинты для групп товаров ==========


class GroupPostDataPhoto(BaseModel):
    """Фото для поста группы"""
    id: int
    file_path: str
    telegram_file_id: str
    item_id: int


class GroupPostDataItem(BaseModel):
    """Товар в группе для поста"""
    id: int
    name: str
    size: Optional[List[str]] = None
    current_price_rub: Optional[Decimal] = None
    item_type: Optional[str] = None


class GroupPostDataResponse(BaseModel):
    """Данные для модалки поста группы"""
    group: dict
    items: List[GroupPostDataItem]
    photos: List[GroupPostDataPhoto]


@router.get("/admin/item-groups/{group_id}/post-data", response_model=GroupPostDataResponse)
async def get_group_post_data(
    group_id: int,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Получить данные для поста группы: товары с ценами и все фото с telegram_file_id"""
    result = await session.execute(select(ItemGroup).where(ItemGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена")

    items_result = await session.execute(
        select(Item).options(joinedload(Item.item_type_rel))
        .where(Item.group_id == group_id)
        .order_by(Item.id)
    )
    items = items_result.unique().scalars().all()
    if not items:
        raise HTTPException(status_code=400, detail="В группе нет товаров")

    ctx = await get_finance_price_context()

    items_list = []
    all_photos = []
    for item in items:
        current_price = await calculate_item_price(item, ctx)
        items_list.append(GroupPostDataItem(
            id=item.id,
            name=item.name,
            size=item.size,
            current_price_rub=current_price,
            item_type=item.item_type_rel.name if item.item_type_rel else None
        ))
        photos_result = await session.execute(
            select(ItemPhoto)
            .where(ItemPhoto.item_id == item.id)
            .where(ItemPhoto.telegram_file_id.isnot(None))
            .order_by(func.coalesce(ItemPhoto.sort_order, 999999).asc(), ItemPhoto.id)
        )
        for photo in photos_result.scalars().all():
            all_photos.append(GroupPostDataPhoto(
                id=photo.id,
                file_path=photo.file_path,
                telegram_file_id=photo.telegram_file_id,
                item_id=item.id
            ))

    return GroupPostDataResponse(
        group={"id": group.id, "name": group.name},
        items=items_list,
        photos=all_photos
    )


@router.post("/admin/item-groups/{group_id}/post-to-telegram")
async def post_group_to_telegram(
    group_id: int,
    message_text: str = Form(...),
    photo_ids: str = Form(...),
    additional_photos: list[UploadFile] = File(default=[]),
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    raise HTTPException(
        status_code=410,
        detail="Публикация групп в Telegram-канал отключена в StoreKPLite.",
    )


@router.get("/admin/item-groups", response_model=List[ItemGroupResponse])
async def list_item_groups(
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Получить список всех групп товаров"""
    result = await session.execute(select(ItemGroup).order_by(ItemGroup.name))
    groups = result.scalars().all()
    
    groups_with_counts = []
    for group in groups:
        items_count_result = await session.execute(
            select(Item).where(Item.group_id == group.id)
        )
        items_count = len(items_count_result.scalars().all())
        groups_with_counts.append(ItemGroupResponse(
            id=group.id,
            name=group.name,
            created_at=group.created_at,
            items_count=items_count
        ))
    
    return groups_with_counts


@router.get("/admin/item-groups/{group_id}", response_model=ItemGroupWithItemsResponse)
async def get_item_group(
    group_id: int,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Получить группу товаров с товарами"""
    result = await session.execute(select(ItemGroup).where(ItemGroup.id == group_id))
    group = result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена")
    
    # Получаем товары в группе
    items_result = await session.execute(
        select(Item).options(joinedload(Item.item_type_rel)).where(Item.group_id == group_id).order_by(Item.id)
    )
    items = items_result.unique().scalars().all()
    
    # Преобразуем товары в словари
    items_list = []
    for item in items:
        photos_result = await session.execute(
            select(ItemPhoto).where(ItemPhoto.item_id == item.id).order_by(
            func.coalesce(ItemPhoto.sort_order, 999999).asc(),
            ItemPhoto.id
        )
        )
        photos = photos_result.scalars().all()
        
        items_list.append({
            "id": item.id,
            "name": item.name,
            "price": float(item.price),
            "item_type": item.item_type_rel.name if item.item_type_rel else None,
            "gender": item.gender,
            "size": item.size,
            "photos_count": len(photos),
            "is_legit": item.is_legit,
        })
    
    return ItemGroupWithItemsResponse(
        id=group.id,
        name=group.name,
        created_at=group.created_at,
        items=items_list
    )


# ---------- Размерные сетки ----------


@router.get("/admin/size-charts", response_model=List[SizeChartListItem])
async def list_size_charts(
    admin=Depends(check_admin_access),
    session: AsyncSession = Depends(get_session),
):
    """Список размерных сеток (для выбора в форме товара)."""
    result = await session.execute(select(SizeChart).order_by(SizeChart.name))
    charts = result.scalars().all()
    return [SizeChartListItem(id=c.id, name=c.name) for c in charts]


@router.get("/admin/size-charts/{chart_id}", response_model=SizeChartResponse)
async def get_size_chart(
    chart_id: int,
    admin=Depends(check_admin_access),
    session: AsyncSession = Depends(get_session),
):
    """Получить размерную сетку по ID."""
    result = await session.execute(select(SizeChart).where(SizeChart.id == chart_id))
    chart = result.scalar_one_or_none()
    if not chart:
        raise HTTPException(status_code=404, detail="Размерная сетка не найдена")
    return SizeChartResponse(id=chart.id, name=chart.name, grid=chart.grid)


@router.post("/admin/size-charts", response_model=SizeChartResponse)
async def create_size_chart(
    request: CreateSizeChartRequest,
    admin=Depends(check_admin_access),
    session: AsyncSession = Depends(get_session),
):
    """Создать размерную сетку."""
    chart = SizeChart(name=request.name.strip(), grid=request.grid)
    session.add(chart)
    await session.commit()
    await session.refresh(chart)
    return SizeChartResponse(id=chart.id, name=chart.name, grid=chart.grid)


@router.patch("/admin/size-charts/{chart_id}", response_model=SizeChartResponse)
async def update_size_chart(
    chart_id: int,
    request: UpdateSizeChartRequest,
    admin=Depends(check_admin_access),
    session: AsyncSession = Depends(get_session),
):
    """Обновить размерную сетку."""
    result = await session.execute(select(SizeChart).where(SizeChart.id == chart_id))
    chart = result.scalar_one_or_none()
    if not chart:
        raise HTTPException(status_code=404, detail="Размерная сетка не найдена")
    if request.name is not None:
        chart.name = request.name.strip()
    if request.grid is not None:
        chart.grid = request.grid
    await session.commit()
    await session.refresh(chart)
    return SizeChartResponse(id=chart.id, name=chart.name, grid=chart.grid)


@router.delete("/admin/size-charts/{chart_id}")
async def delete_size_chart(
    chart_id: int,
    admin=Depends(check_admin_access),
    session: AsyncSession = Depends(get_session),
):
    """Удалить размерную сетку. У товаров size_chart_id станет NULL."""
    result = await session.execute(select(SizeChart).where(SizeChart.id == chart_id))
    chart = result.scalar_one_or_none()
    if not chart:
        raise HTTPException(status_code=404, detail="Размерная сетка не найдена")
    await session.delete(chart)
    await session.commit()
    return {"message": "Размерная сетка удалена", "id": chart_id}


@router.post("/admin/item-groups", response_model=ItemGroupResponse)
async def create_item_group(
    request: CreateItemGroupRequest,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Создать новую группу товаров"""
    group = ItemGroup(name=request.name)
    session.add(group)
    await session.commit()
    await session.refresh(group)
    
    return ItemGroupResponse(
        id=group.id,
        name=group.name,
        created_at=group.created_at,
        items_count=0
    )


@router.put("/admin/item-groups/{group_id}", response_model=ItemGroupResponse)
async def update_item_group(
    group_id: int,
    request: UpdateItemGroupRequest,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Обновить группу товаров"""
    result = await session.execute(select(ItemGroup).where(ItemGroup.id == group_id))
    group = result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена")
    
    if request.name is not None:
        group.name = request.name
    
    await session.commit()
    await session.refresh(group)
    
    # Подсчитываем количество товаров
    items_count_result = await session.execute(
        select(Item).where(Item.group_id == group.id)
    )
    items_count = len(items_count_result.scalars().all())
    
    return ItemGroupResponse(
        id=group.id,
        name=group.name,
        created_at=group.created_at,
        items_count=items_count
    )


@router.delete("/admin/item-groups/{group_id}")
async def delete_item_group(
    group_id: int,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Удалить группу товаров (товары останутся, но group_id станет NULL)"""
    result = await session.execute(select(ItemGroup).where(ItemGroup.id == group_id))
    group = result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена")
    
    # Сначала обнуляем group_id у всех товаров в группе
    await session.execute(
        update(Item)
        .where(Item.group_id == group_id)
        .values(group_id=None)
    )
    
    # Затем удаляем группу
    await session.delete(group)
    await session.commit()
    
    return {"success": True, "message": "Группа удалена"}


@router.post("/admin/item-groups/{group_id}/items")
async def add_items_to_group(
    group_id: int,
    request: AddItemsToGroupRequest,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Добавить товары в группу"""
    # Проверяем существование группы
    group_result = await session.execute(select(ItemGroup).where(ItemGroup.id == group_id))
    group = group_result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена")
    
    # Проверяем, что товары существуют и не находятся ни в какой группе
    added_count = 0
    errors = []

    for item_id in request.item_ids:
        item_result = await session.execute(select(Item).where(Item.id == item_id))
        item = item_result.scalar_one_or_none()

        if not item:
            errors.append(f"Товар {item_id} не найден")
            continue

        # Проверяем, что товар не состоит ни в какой группе (включая текущую)
        if item.group_id is not None:
            if item.group_id == group_id:
                errors.append(f"Товар {item_id} уже находится в этой группе")
            else:
                errors.append(f"Товар {item_id} уже находится в другой группе (ID: {item.group_id})")
            continue

        item.group_id = group_id
        added_count += 1
    
    await session.commit()
    
    return {
        "success": True,
        "added_count": added_count,
        "errors": errors if errors else None
    }


@router.delete("/admin/item-groups/{group_id}/items/{item_id}")
async def remove_item_from_group(
    group_id: int,
    item_id: int,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Удалить товар из группы"""
    result = await session.execute(
        select(Item).where(Item.id == item_id).where(Item.group_id == group_id)
    )
    item = result.scalar_one_or_none()
    
    if not item:
        raise HTTPException(status_code=404, detail="Товар не найден в этой группе")
    
    item.group_id = None
    await session.commit()
    
    return {"success": True, "message": "Товар удален из группы"}


class UpdateGroupItemsPriceRequest(BaseModel):
    price: Decimal


class UpdateGroupItemsServiceFeeRequest(BaseModel):
    service_fee_percent: Decimal


@router.put("/admin/item-groups/{group_id}/items/price", response_model=dict)
async def update_group_items_price(
    group_id: int,
    request: UpdateGroupItemsPriceRequest,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Массово изменить цену для всех товаров в группе"""
    # Проверяем существование группы
    group_result = await session.execute(select(ItemGroup).where(ItemGroup.id == group_id))
    group = group_result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена")
    
    # Получаем все товары в группе
    items_result = await session.execute(select(Item).where(Item.group_id == group_id))
    items = items_result.scalars().all()
    
    if not items:
        raise HTTPException(status_code=400, detail="В группе нет товаров")
    
    updated_count = 0
    
    # Обновляем цену для всех товаров в группе
    for item in items:
        item.price = request.price
        updated_count += 1
    
    await session.commit()
    
    ctx = await get_finance_price_context()
    from api.products.routers.price_history import upsert_price_history_4h_bucket
    for item in items:
        new_price_rub = await calculate_item_price(item, ctx)
        await upsert_price_history_4h_bucket(session, item.id, new_price_rub)
    await session.commit()
    
    return {
        "success": True,
        "updated_count": updated_count,
        "message": f"Цена обновлена для {updated_count} товаров в группе"
    }


@router.put("/admin/item-groups/{group_id}/items/service-fee", response_model=dict)
async def update_group_items_service_fee(
    group_id: int,
    request: UpdateGroupItemsServiceFeeRequest,
    admin = Depends(check_admin_access),
    session: AsyncSession = Depends(get_session)
):
    """Массово изменить сервисный сбор для всех товаров в группе"""
    # Проверяем существование группы
    group_result = await session.execute(select(ItemGroup).where(ItemGroup.id == group_id))
    group = group_result.scalar_one_or_none()
    
    if not group:
        raise HTTPException(status_code=404, detail="Группа не найдена")
    
    # Получаем все товары в группе
    items_result = await session.execute(select(Item).where(Item.group_id == group_id))
    items = items_result.scalars().all()
    
    if not items:
        raise HTTPException(status_code=400, detail="В группе нет товаров")
    
    updated_count = 0
    
    # Обновляем сервисный сбор для всех товаров в группе
    for item in items:
        item.service_fee_percent = request.service_fee_percent
        updated_count += 1
    
    await session.commit()
    
    ctx = await get_finance_price_context()
    from api.products.routers.price_history import upsert_price_history_4h_bucket
    for item in items:
        new_price_rub = await calculate_item_price(item, ctx)
        await upsert_price_history_4h_bucket(session, item.id, new_price_rub)
    await session.commit()
    
    return {
        "success": True,
        "updated_count": updated_count,
        "message": f"Сервисный сбор обновлен для {updated_count} товаров в группе"
    }


# ========== Админские эндпоинты для заказов ==========

from api.products.models.order import Order
from api.products.routers.orders import OrderResponse
from api.products.utils.order_helpers import (
    cdek_delivery_calc_insurance_extras,
    compute_order_total,
    compute_order_amount_due,
    delivery_cost_from_order_snapshot,
)
from api.products.utils.tryon_order_discount import complete_tryon_for_order, release_tryon_for_order
from api.products.utils.order_promo_display import (
    admin_order_data_for_response,
    system_photo_promo_lines_by_order,
    order_data_with_system_promo_flags,
)
from datetime import datetime
from decimal import Decimal as DecimalType


def _order_total_payable(order: Order, *, exclude_returned: bool = False) -> float:
    """Итог к оплате по заказу: товары - скидка за примерки + доставка."""
    return compute_order_amount_due(
        order.order_data,
        float(getattr(order, "tryon_discount_rub", 0) or 0),
        delivery_cost_from_order_snapshot(order.order_data),
        exclude_returned=exclude_returned,
    )


async def send_telegram_message(tgid: Optional[int], text: str) -> None:
    """
    Отправка сообщения пользователю в Telegram (API_BASE_URL/bot-api или прямой BOT_TOKEN).
    Используется для уведомлений о статусах заказов.
    """
    if not tgid:
        return

    if not _tg_bridge.telegram_outbound_configured():
        logger.warning("Исходящий Telegram не настроен (API_BASE_URL+bot-api или BOT_TOKEN), уведомление не отправлено")
        return

    await _tg_bridge.telegram_send_message(tgid, text, timeout=10.0)


async def send_vk_message(vkid: Optional[int], text: str) -> None:
    """
    Отправка сообщения пользователю в VK через Bot API.
    Используется для уведомлений о статусах заказов.
    """
    if not vkid:
        return
    
    vk_bot_token = getenv("VK_BOT_TOKEN")
    if not vk_bot_token:
        logger.warning("VK_BOT_TOKEN не указан, уведомление пользователю не отправлено")
        return
    
    vk_api_url = "https://api.vk.com/method"
    
    try:
        import random
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.post(
                f"{vk_api_url}/messages.send",
                params={
                    "access_token": vk_bot_token,
                    "user_id": vkid,
                    "message": text,
                    "random_id": random.randint(1, 2**31),
                    "v": "5.131"
                }
            )
            response.raise_for_status()
            result = response.json()
            
            if "error" in result:
                logger.error(f"VK API вернул ошибку при отправке уведомления пользователю {vkid}: {result['error']}")
            else:
                logger.info(f"Уведомление отправлено пользователю VK {vkid}")
    except httpx.HTTPStatusError as e:
        logger.error(f"Ошибка HTTP при отправке уведомления пользователю VK {vkid}: {e.response.status_code} - {e.response.text}")
    except Exception as e:
        logger.error(f"Ошибка при отправке уведомления пользователю VK {vkid}: {e}", exc_info=True)


class DeliveryInfoForAdmin(BaseModel):
    """Информация о доставке для админки"""
    delivery_status_id: Optional[int] = None
    delivery_status_name: Optional[str] = None
    additional_info: Optional[str] = None


class AdminOrderResponse(OrderResponse):
    """Расширенный ответ для админов с дополнительной информацией о пользователе"""
    cancel_reason: Optional[str] = None
    refund_on_cancel: Optional[bool] = None
    user_tgid: Optional[int] = None
    user_vkid: Optional[int] = None
    delivery: Optional[DeliveryInfoForAdmin] = None
    phone_number: Optional[str] = None
    delivery_id: Optional[int] = None
    # Контактные данные пользователя (для связи)
    user_firstname: Optional[str] = None
    user_username: Optional[str] = None
    user_email: Optional[str] = None
    user_phone: Optional[str] = None  # номер из профиля (country_code + phone_local)


class AdminOrderListResponse(BaseModel):
    """Постраничный список заказов для админки (как каталог: total + has_more)."""
    items: List[AdminOrderResponse]
    total: int
    has_more: bool


class UpdateOrderStatusRequest(BaseModel):
    new_status: str
    cancel_reason: Optional[str] = None  # Причина отмены (обязательна при отмене заказа с привязанным пользователем)


class UpdatePaidAmountRequest(BaseModel):
    paid_amount: float


class ManualOrderItemRequest(BaseModel):
    """Товар для ручного создания заказа"""
    item_id: Optional[int] = None  # ID товара из каталога (если None - кастомный товар)
    name: str
    size: Optional[str] = None
    chinese_name: Optional[str] = None  # для кастомной позиции — опционально, в order_data.items
    quantity: int
    price: float  # Цена в рублях
    link: Optional[str] = None  # Ссылка на товар (для кастомных товаров)
    # Путь после POST /admin/orders/manual/upload-custom-photo (только для кастомной позиции)
    photo: Optional[str] = None
    # Для кастомной позиции — обязательны; для каталога игнорируются (берутся из карточки товара)
    estimated_weight_kg: Optional[float] = None
    length_cm: Optional[int] = None
    width_cm: Optional[int] = None
    height_cm: Optional[int] = None


class ManualDeliveryPreviewItem(BaseModel):
    """Позиция для предпросчёта доставки (ручной заказ до создания)."""

    item_id: Optional[int] = None
    name: str = ""
    quantity: int = 1
    price: float = 0.0
    estimated_weight_kg: Optional[float] = None
    length_cm: Optional[int] = None
    width_cm: Optional[int] = None
    height_cm: Optional[int] = None


class ManualDeliveryPreviewRequest(BaseModel):
    """Предпросчёт стоимости доставки тем же POST /calculate-cost, что и при чекауте."""

    items: List[ManualDeliveryPreviewItem]
    delivery_method_code: str = "CDEK"
    delivery_city_code: int
    cdek_delivery_point_code: str


class ManualDeliveryPreviewResponse(BaseModel):
    delivery_cost_rub: Optional[float] = None
    cdek_tariff_code: Optional[int] = None
    cdek_delivery_sum_base_rub: Optional[float] = None
    cdek_total_sum_rub: Optional[float] = None


class ManualOrderDeliveryInput(BaseModel):
    """Доставка при ручном создании заказа (снимок в order_data.delivery_snapshot)."""

    delivery_method_code: str
    recipient_name: Optional[str] = None
    address: Optional[str] = None
    postal_code: Optional[str] = None
    delivery_city_code: Optional[int] = None
    delivery_city: Optional[str] = None
    cdek_delivery_point_code: Optional[str] = None
    delivery_cost_rub: Optional[float] = None
    local_pickup_point_id: Optional[int] = None


class CreateManualOrderRequest(BaseModel):
    """Запрос на ручное создание заказа"""
    items: List[ManualOrderItemRequest]
    user_id: int  # Внутренний id пользователя (обязательная привязка)
    phone_number: str
    is_paid: Optional[bool] = False  # Если True, заказ создается с полной оплатой
    delivery: Optional[ManualOrderDeliveryInput] = None


class ItemCatalogMeasurementUpdate(BaseModel):
    """Правки веса/габаритов по товару каталога перед созданием накладной — сохраняются в items."""

    item_id: int
    estimated_weight_kg: Optional[float] = None
    length_cm: Optional[int] = None
    width_cm: Optional[int] = None
    height_cm: Optional[int] = None


class OrderLineParcelUpdate(BaseModel):
    """Правки веса/габаритов по строке заказа (в т.ч. кастом) — сохраняются в order_data.items[line_index]."""

    line_index: int
    estimated_weight_kg: float
    length_cm: int
    width_cm: int
    height_cm: int


class CreateShipmentRequest(BaseModel):
    """Опциональные переопределения из админки. Если не переданы — берутся из заказа (delivery_snapshot при чекауте)."""
    recipient_name: Optional[str] = None
    address: Optional[str] = None
    city_code: Optional[int] = None
    postal_code: Optional[str] = None
    delivery_price_rub: Optional[float] = None
    cdek_delivery_point_code: Optional[str] = None  # код ПВЗ/офиса СДЭК → delivery_point в API СДЭК
    item_catalog_updates: Optional[List[ItemCatalogMeasurementUpdate]] = None
    order_line_updates: Optional[List[OrderLineParcelUpdate]] = None


class CreateShipmentResponse(BaseModel):
    """Ответ создания накладной: заказ и при наличии — PDF накладной в base64."""
    order: AdminOrderResponse
    shipment_pdf_base64: Optional[str] = None
    cdek_order_uuid: Optional[str] = None
    cdek_delivery_sum_rub: Optional[float] = None


class ShipmentCdekPollResponse(BaseModel):
    """Long-poll статуса накладной СДЭК: обновлённый заказ и при готовности — PDF (base64)."""
    pending: bool
    cdek_invalid: bool = False
    cdek_detail: Optional[str] = None
    shipment_pdf_base64: Optional[str] = None
    cdek_delivery_sum_rub: Optional[float] = None
    order: AdminOrderResponse


class ShipmentItemPreviewLine(BaseModel):
    """Строка заказа для модалки перед отправкой в СДЭК."""

    line_index: int
    item_id: Optional[int] = None
    name: str
    quantity: int
    estimated_weight_kg: Optional[float] = None
    length_cm: Optional[int] = None
    width_cm: Optional[int] = None
    height_cm: Optional[int] = None
    weight_gram_per_unit: int
    catalog_editable: bool = False
    order_line_editable: bool = False


class ShipmentParcelSummary(BaseModel):
    weight_gram: int
    length_cm: int
    width_cm: int
    height_cm: int
    total_weight_kg: float
    total_volume_cm3: int


class ShipmentItemsPreviewResponse(BaseModel):
    lines: List[ShipmentItemPreviewLine]
    parcel: ShipmentParcelSummary


def _shipment_preview_int_or_none(val: Any) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


async def shipment_items_preview(
    order_id: int,
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session),
):
    raise HTTPException(
        status_code=410,
        detail="Функционал накладных/автосборки отключен в StoreKPLite.",
    )


# Значение «трек» от delivery при отсутствии cdek_number в ответе СДЭК — UUID заказа API, не номер накладной.
_TRACKING_IS_CDEK_ORDER_UUID = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}\Z",
    re.I,
)


def _tracking_looks_like_cdek_api_uuid(val: Any) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return bool(s) and bool(_TRACKING_IS_CDEK_ORDER_UUID.match(s))


def _apply_delivery_carrier_settlement_cdek_from_api(order: Order, amount: float) -> bool:
    """
    Записывает факт delivery_sum СДЭК в order_data.delivery_carrier_settlement (блок «Сумма к оплате перевозчику»,
    агрегат в финансах). Месяц начисления: уже заданный в учёте по заказу, иначе месяц created_at заказа.
    """
    from datetime import datetime, timezone

    if amount <= 0:
        return False
    od = dict(order.order_data or {})
    month: Optional[str] = None
    prev = od.get("delivery_carrier_settlement")
    if isinstance(prev, dict):
        m = prev.get("accrual_month")
        if isinstance(m, str) and m.strip():
            month = m.strip()
    if not month and order.created_at:
        month = order.created_at.strftime("%Y-%m")
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    od["delivery_carrier_settlement"] = {
        "payable_rub": round(float(amount), 2),
        "carrier": "CDEK",
        "accrual_month": month,
        "note": "Авто: фактическая стоимость доставки СДЭК (delivery_sum) после накладной",
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    order.order_data = od
    flag_modified(order, "order_data")
    return True


async def create_order_shipment(
    order_id: int,
    request: Optional[CreateShipmentRequest] = Body(None),
    admin = Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    raise HTTPException(
        status_code=410,
        detail="Автосоздание CDEK-накладных отключено в StoreKPLite.",
    )


async def poll_shipment_cdek_status(
    order_id: int,
    wait_seconds: int = Query(60, ge=5, le=120),
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session),
):
    raise HTTPException(
        status_code=410,
        detail="Опрос CDEK-накладных отключен в StoreKPLite.",
    )


async def get_order_waybill_pdf(
    order_id: int,
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session),
):
    raise HTTPException(
        status_code=410,
        detail="Получение CDEK-накладных отключено в StoreKPLite.",
    )


_TERMINAL_ORDER_STATUSES = ("отменен", "завершен")


async def _fetch_active_order_to_delivery_map() -> Dict[int, int]:
    """Поставки отключены в StoreKPLite."""
    return {}


@router.get("/admin/orders", response_model=AdminOrderListResponse)
async def list_orders(
    status_filter: Optional[str] = Query(None, description="Фильтр по статусу"),
    search: Optional[str] = Query(None, description="Поиск по номеру заказа (целое id)"),
    delivery_filter: Optional[str] = Query(
        None,
        description="Фильтр по поставке: in_delivery — в активной поставке, free — не в поставке",
    ),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=1000),
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session),
):
    """
    Список заказов для админки с пагинацией.

    Заказы в статусах «отменен» и «завершен» не попадают в выборку по умолчанию;
    они возвращаются при явном фильтре по статусу, фильтре по поставке,
    или при поиске по числовому id заказа.
    """
    df = (delivery_filter or "").strip()
    if df and df not in ("in_delivery", "free"):
        raise HTTPException(
            status_code=400,
            detail="delivery_filter должен быть пустым, in_delivery или free",
        )

    delivery_map = await _fetch_active_order_to_delivery_map()
    in_delivery_ids = list(delivery_map.keys())

    clauses: List[Any] = []
    if status_filter and status_filter.strip():
        clauses.append(Order.status == status_filter.strip())

    search_stripped = (search or "").strip()
    order_id_from_search: Optional[int] = None
    if search_stripped:
        try:
            order_id_from_search = int(search_stripped)
        except ValueError:
            pass
    if order_id_from_search is not None:
        clauses.append(Order.id == order_id_from_search)

    if df == "in_delivery":
        if not in_delivery_ids:
            return AdminOrderListResponse(items=[], total=0, has_more=False)
        clauses.append(Order.id.in_(in_delivery_ids))
    elif df == "free" and in_delivery_ids:
        clauses.append(Order.id.notin_(in_delivery_ids))

    has_status_filter = bool(status_filter and status_filter.strip())
    has_delivery_filter = bool(df)
    has_numeric_id_search = order_id_from_search is not None
    exclude_terminal = not has_status_filter and not has_delivery_filter and not has_numeric_id_search
    if exclude_terminal:
        clauses.append(~Order.status.in_(_TERMINAL_ORDER_STATUSES))

    where_clause = and_(*clauses) if clauses else true()

    count_stmt = select(func.count()).select_from(Order).where(where_clause)
    total = int(await session.scalar(count_stmt) or 0)

    list_stmt = (
        select(Order)
        .options(joinedload(Order.delivery).joinedload(OrderDelivery.delivery_status))
        .where(where_clause)
        .order_by(Order.created_at.desc())
        .offset(skip)
        .limit(limit)
    )
    result = await session.execute(list_stmt)
    orders = result.scalars().unique().all()
    sys_promo_map = await system_photo_promo_lines_by_order(session, [o.id for o in orders])

    unique_user_ids = sorted({o.user_id for o in orders if o.user_id})
    user_by_id: Dict[int, Optional[Dict[str, Any]]] = {}
    if unique_user_ids:
        try:
            async with httpx.AsyncClient() as client:
                r = await client.post(
                    f"{USERS_SERVICE_URL.rstrip('/')}/internal/users/by-ids",
                    headers={"X-Internal-Token": INTERNAL_TOKEN},
                    json={"user_ids": unique_user_ids},
                    timeout=15.0,
                )
                if r.status_code == 200:
                    for u in (r.json() or {}).get("users") or []:
                        uid = u.get("id")
                        if isinstance(uid, int):
                            user_by_id[uid] = u
                else:
                    logger.warning(
                        "Пакетный запрос профилей пользователей: HTTP %s %s",
                        r.status_code,
                        (r.text or "")[:200],
                    )
        except Exception as e:
            logger.warning("Не удалось получить профили пользователей пакетом: %s", e)

    orders_with_users: List[AdminOrderResponse] = []
    for order in orders:
        user_tgid = None
        user_vkid = None
        user_firstname = None
        user_username = None
        user_email = None
        user_phone = None
        if order.user_id:
            user_data = user_by_id.get(order.user_id)
            if user_data:
                user_tgid = user_data.get("tgid")
                user_vkid = user_data.get("vkid")
                user_firstname = user_data.get("firstname")
                user_username = user_data.get("username")
                user_email = user_data.get("email")
                cc = (user_data.get("country_code") or "").strip()
                pl = (user_data.get("phone_local") or "").strip()
                user_phone = (cc + " " + pl).strip() if (cc or pl) else None

        delivery_info = None
        if order.delivery:
            delivery_info = DeliveryInfoForAdmin(
                delivery_status_id=order.delivery.delivery_status_id,
                delivery_status_name=order.delivery.delivery_status.name if order.delivery.delivery_status else None,
                additional_info=order.delivery.additional_info,
            )

        did = delivery_map.get(order.id)
        orders_with_users.append(
            AdminOrderResponse(
                id=order.id,
                user_id=order.user_id,
                order_data=order_data_with_system_promo_flags(order.order_data, sys_promo_map.get(order.id, set())),
                status=order.status,
                created_at=order.created_at,
                updated_at=order.updated_at,
                paid_amount=float(order.paid_amount),
                order_total=_order_total_payable(order),
                tracking_number=getattr(order, "tracking_number", None),
                cancel_reason=order.cancel_reason,
                refund_on_cancel=order.refund_on_cancel,
                user_tgid=user_tgid,
                user_vkid=user_vkid,
                delivery=delivery_info,
                phone_number=order.phone_number,
                delivery_id=did,
                is_from_stock=order.is_from_stock,
                recipient_name=getattr(order, "recipient_name", None),
                user_firstname=user_firstname,
                user_username=user_username,
                user_email=user_email,
                user_phone=user_phone,
            )
        )

    loaded = len(orders_with_users)
    return AdminOrderListResponse(
        items=orders_with_users,
        total=total,
        has_more=(skip + loaded) < total,
    )


class ExportOrdersExcelRequest(BaseModel):
    order_ids: List[int]


def _resolve_photo_path(file_path: str) -> Optional[Path]:
    """Возвращает абсолютный путь к файлу фото или None если файл не найден."""
    if not file_path or not file_path.strip():
        return None
    p = Path(file_path.strip())
    if p.is_absolute() and p.exists():
        return p
    if p.exists():
        return p
    # Часто в БД хранится uploads/items/xxx.jpg или только xxx.jpg
    if "uploads" in file_path.replace("\\", "/"):
        p2 = Path(file_path.replace("\\", "/"))
        if p2.exists():
            return p2
    name = p.name
    candidate = UPLOAD_DIR / name
    if candidate.exists():
        return candidate
    return None


async def _get_orders_export_rows(session: AsyncSession, order_ids: List[int]) -> List[Dict[str, Any]]:
    """Общая выборка для экспорта заказов: сводка (товар+размер) -> name/chinese_name, link, size, quantity, photo_path."""
    result = await session.execute(
        select(Order).where(Order.id.in_(order_ids)).order_by(Order.id)
    )
    orders = result.scalars().unique().all()
    if not orders:
        return []
    aggregated: Dict[tuple, Dict[str, Any]] = {}
    for order in orders:
        items = (order.order_data or {}).get("items") or []
        for it in items:
            item_id = it.get("item_id")
            name = (it.get("name") or "").strip()
            link = it.get("link") or ""
            size = (it.get("size") or "").strip()
            quantity = int(it.get("quantity") or 1)
            key = (item_id, size)
            if key not in aggregated:
                aggregated[key] = {
                    "item_id": item_id,
                    "name": name,
                    "chinese_name": normalize_optional_name(it.get("chinese_name")),
                    "link": link,
                    "size": size,
                    "quantity": 0,
                }
            aggregated[key]["quantity"] += quantity
    if not aggregated:
        return []
    item_ids = list(dict.fromkeys([v["item_id"] for v in aggregated.values() if v["item_id"] is not None]))
    first_photo_by_item: Dict[int, str] = {}
    chinese_name_by_item: Dict[int, Optional[str]] = {}
    if item_ids:
        items_result = await session.execute(select(Item.id, Item.chinese_name).where(Item.id.in_(item_ids)))
        for item_id, chinese_name in items_result.all():
            chinese_name_by_item[item_id] = normalize_optional_name(chinese_name)
        photos_result = await session.execute(
            select(ItemPhoto)
            .where(ItemPhoto.item_id.in_(item_ids))
            .order_by(ItemPhoto.item_id, func.coalesce(ItemPhoto.sort_order, 999999).asc(), ItemPhoto.id)
        )
        for p in photos_result.scalars().all():
            if p.item_id not in first_photo_by_item and p.file_path:
                first_photo_by_item[p.item_id] = p.file_path
    rows = []
    for (item_id, size), data in sorted(aggregated.items(), key=lambda x: (x[0][0] or 0, x[0][1])):
        rows.append({
            "photo_path": first_photo_by_item.get(item_id) if item_id is not None else None,
            "name": data["name"],
            "chinese_name": data.get("chinese_name") or (chinese_name_by_item.get(item_id) if item_id is not None else None),
            "link": data["link"],
            "size": data["size"],
            "quantity": data["quantity"],
        })
    return rows


@router.post("/admin/orders/export-excel")
async def export_orders_excel(
    request: ExportOrdersExcelRequest,
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session),
):
    """Экспорт выбранных заказов в Excel для партнёра: сводка — какой товар, какой размер, сколько штук заказать (фото, название, ссылка, размер, количество)."""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="Модуль openpyxl не установлен. Установите: pip install openpyxl")

    if not request.order_ids:
        raise HTTPException(status_code=400, detail="Укажите хотя бы один заказ (order_ids)")

    rows = await _get_orders_export_rows(session, request.order_ids)
    if not rows:
        raise HTTPException(status_code=400, detail="В выбранных заказах нет позиций")

    wb = Workbook()
    ws = wb.active
    ws.title = "Заказ партнёру"
    ws.append(["Фото", "Название", "Название (CN)", "Ссылка", "Размер", "Кол-во"])

    IMG_HEIGHT_PX = 90
    IMG_WIDTH_PX = 120
    ROW_HEIGHT_POINTS = 68
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10

    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in (1, 2, 3, 4, 5, 6):
        ws.cell(row=1, column=col).alignment = alignment_center

    for row_idx, r in enumerate(rows, start=2):
        ws.row_dimensions[row_idx].height = ROW_HEIGHT_POINTS
        ws.cell(row=row_idx, column=2, value=r["name"]).alignment = alignment_center
        ws.cell(row=row_idx, column=3, value=r.get("chinese_name") or "").alignment = alignment_center
        link_cell = ws.cell(row=row_idx, column=4, value=r["link"] or "")
        link_cell.alignment = alignment_center
        if r.get("link") and str(r["link"]).strip():
            link_cell.hyperlink = r["link"].strip()
        ws.cell(row=row_idx, column=5, value=r["size"]).alignment = alignment_center
        ws.cell(row=row_idx, column=6, value=r["quantity"]).alignment = alignment_center

        photo_path_str = r.get("photo_path")
        if photo_path_str:
            resolved = _resolve_photo_path(photo_path_str)
            if resolved and resolved.is_file():
                try:
                    img = XLImage(str(resolved))
                    h, w = img.height, img.width
                    if h > 0:
                        scale = IMG_HEIGHT_PX / h
                        img.width = min(int(w * scale), IMG_WIDTH_PX)
                        img.height = IMG_HEIGHT_PX
                    img.anchor = f"A{row_idx}"
                    ws.add_image(img)
                except Exception as e:
                    logger.warning("Не удалось вставить изображение %s: %s", resolved, e)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _register_pdf_cyrillic_font():
    """Регистрирует шрифт с поддержкой кириллицы для ReportLab. Возвращает имя шрифта или None."""
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return None
    font_name = "CyrillicPDF"
    try:
        if pdfmetrics.getFont(font_name) is not None:
            return font_name
    except (KeyError, Exception):
        pass
    candidates = []
    if os.name == "nt":
        windir = os.environ.get("WINDIR", "C:\\Windows")
        candidates.append(os.path.join(windir, "Fonts", "arial.ttf"))
        candidates.append(os.path.join(windir, "Fonts", "Arial.ttf"))
    else:
        candidates.extend([
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/TTF/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        ])
    for path in candidates:
        if path and os.path.isfile(path):
            try:
                pdfmetrics.registerFont(TTFont(font_name, path))
                return font_name
            except Exception as e:
                logger.debug("Не удалось загрузить шрифт %s: %s", path, e)
    return None


@router.post("/admin/orders/export-pdf")
async def export_orders_pdf(
    request: ExportOrdersExcelRequest,
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session),
):
    """Экспорт заказов в PDF (фото отображаются на мобильных и в Google). Та же сводка, что и в Excel."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image as RLImage, Spacer
    except ImportError:
        raise HTTPException(status_code=500, detail="Модуль reportlab не установлен. Установите: pip install reportlab")

    if not request.order_ids:
        raise HTTPException(status_code=400, detail="Укажите хотя бы один заказ (order_ids)")

    rows_data = await _get_orders_export_rows(session, request.order_ids)
    if not rows_data:
        raise HTTPException(status_code=400, detail="В выбранных заказах нет позиций")

    styles = getSampleStyleSheet()
    cyrillic_font = _register_pdf_cyrillic_font()
    if cyrillic_font:
        from reportlab.lib.styles import ParagraphStyle
        for key in ("Normal", "Heading1", "Heading2"):
            base = styles[key]
            styles.add(ParagraphStyle(
                name=f"{key}Cyrillic",
                fontName=cyrillic_font,
                fontSize=base.fontSize,
                leading=base.leading,
                spaceBefore=base.spaceBefore,
                spaceAfter=base.spaceAfter,
                alignment=base.alignment,
            ))
        style = styles["NormalCyrillic"]
    else:
        style = styles["Normal"]
    img_w, img_h = 30 * mm, 25 * mm
    table_data = [[
        Paragraph("<b>Фото</b>", style),
        Paragraph("<b>Название</b>", style),
        Paragraph("<b>Название (CN)</b>", style),
        Paragraph("<b>Ссылка</b>", style),
        Paragraph("<b>Размер</b>", style),
        Paragraph("<b>Кол-во</b>", style),
    ]]

    import tempfile
    temp_dir = tempfile.mkdtemp()
    try:
        for r in rows_data:
            resolved = _resolve_photo_path(r.get("photo_path") or "") if r.get("photo_path") else None
            if resolved and resolved.is_file():
                try:
                    pil_img = Image.open(resolved)
                    if pil_img.mode in ("RGBA", "P"):
                        pil_img = pil_img.convert("RGB")
                    # Ресайз с запасом по разрешению (~2.5× от размера ячейки), чтобы PDF ~2–3 МБ на 7 фото
                    target_w = min(int(img_w * 2.5), pil_img.width)
                    target_h = min(int(img_h * 2.5), pil_img.height)
                    pil_img.thumbnail((target_w, target_h), Image.Resampling.LANCZOS)
                    fd, path = tempfile.mkstemp(suffix=".jpg", dir=temp_dir)
                    pil_img.save(path, format="JPEG", quality=90, optimize=True)
                    os.close(fd)
                    img = RLImage(path, width=img_w, height=img_h)
                except Exception:
                    img = Paragraph("—", style)
            else:
                img = Paragraph("—", style)
            name = (r.get("name") or "").replace("&", "&amp;").replace("<", "&lt;")
            chinese_name = ((r.get("chinese_name") or "").replace("&", "&amp;").replace("<", "&lt;"))[:200]
            raw_link = (r.get("link") or "").strip()
            link_escaped = raw_link.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
            link_display = (raw_link[:150] + ("…" if len(raw_link) > 150 else "")).replace("&", "&amp;").replace("<", "&lt;")
            link_cell = Paragraph(
                f'<a href="{link_escaped}" color="blue">{link_display}</a>' if raw_link else link_display,
                style
            )
            table_data.append([
                img,
                Paragraph(name[:200] + ("…" if len(name) > 200 else ""), style),
                Paragraph(chinese_name, style),
                link_cell,
                Paragraph(str(r.get("size") or ""), style),
                Paragraph(str(r.get("quantity") or 0), style),
            ])

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
        col_widths = [img_w + 3 * mm, 40 * mm, 30 * mm, 45 * mm, 14 * mm, 14 * mm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ]))
        doc.build([table])
        buf.seek(0)
        filename = f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        return Response(
            content=buf.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        import shutil
        if temp_dir and os.path.isdir(temp_dir):
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
            except Exception:
                pass


@router.get("/admin/orders/{order_id}", response_model=AdminOrderResponse)
async def get_order_by_id(
    order_id: int,
    admin = Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    """Получить детали заказа (для админов)"""
    result = await session.execute(
        select(Order)
        .options(joinedload(Order.delivery).joinedload(OrderDelivery.delivery_status))
        .where(Order.id == order_id)
    )
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    
    user_tgid = None
    user_vkid = None
    user_firstname = None
    user_username = None
    user_email = None
    user_phone = None
    if order.user_id:
        users_service_url = getenv("USERS_SERVICE_URL", "http://users-service:8001")
        try:
            async with httpx.AsyncClient() as client:
                user_response = await client.get(
                    f"{users_service_url}/users/{order.user_id}",
                    timeout=5.0
                )
                if user_response.status_code == 200:
                    user_data = user_response.json()
                    user_tgid = user_data.get("tgid")
                    user_vkid = user_data.get("vkid")
                    user_firstname = user_data.get("firstname")
                    user_username = user_data.get("username")
                    user_email = user_data.get("email")
                    cc = (user_data.get("country_code") or "").strip()
                    pl = (user_data.get("phone_local") or "").strip()
                    user_phone = (cc + " " + pl).strip() if (cc or pl) else None
        except Exception as e:
            logger.warning(f"Не удалось получить информацию о пользователе {order.user_id}: {e}")
    
    delivery_info = None
    if order.delivery:
        delivery_info = DeliveryInfoForAdmin(
            delivery_status_id=order.delivery.delivery_status_id,
            delivery_status_name=order.delivery.delivery_status.name if order.delivery.delivery_status else None,
            additional_info=order.delivery.additional_info
        )
    
    return AdminOrderResponse(
        id=order.id,
        user_id=order.user_id,
        order_data=await admin_order_data_for_response(session, order_id=order.id, order_data=order.order_data),
        status=order.status,
        created_at=order.created_at,
        updated_at=order.updated_at,
        paid_amount=float(order.paid_amount),
        order_total=_order_total_payable(order),
        tracking_number=getattr(order, "tracking_number", None),
        cancel_reason=order.cancel_reason,
        refund_on_cancel=order.refund_on_cancel,
        user_tgid=user_tgid,
        user_vkid=user_vkid,
        delivery=delivery_info,
        phone_number=order.phone_number,
        is_from_stock=order.is_from_stock,
        recipient_name=getattr(order, "recipient_name", None),
        user_firstname=user_firstname,
        user_username=user_username,
        user_email=user_email,
        user_phone=user_phone,
    )


@router.post("/admin/orders/{order_id}/status", response_model=AdminOrderResponse)
async def update_order_status(
    order_id: int,
    request: UpdateOrderStatusRequest,
    admin = Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    """Обновить статус заказа (для админов)"""
    # Проверяем валидность статуса
    valid_statuses = ["Ожидает", "Выкуп", "в работе", "Собран", "отменен", "завершен"]
    if request.new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail="Неверный статус")
    
    # Загружаем заказ с предзагруженными связанными объектами delivery и delivery_status
    # чтобы избежать lazy loading вне асинхронного контекста
    result = await session.execute(
        select(Order)
        .options(joinedload(Order.delivery).joinedload(OrderDelivery.delivery_status))
        .where(Order.id == order_id)
    )
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    
    # Сохраняем ссылку на класс DeliveryInfoForAdmin для использования в функции
    # (чтобы избежать проблем с областью видимости)
    DeliveryInfoForAdminClass = DeliveryInfoForAdmin
    
    old_status = order.status
    
    # Если статус не меняется
    if old_status == request.new_status:
        # Возвращаем текущий заказ
        user_tgid = None
        user_vkid = None
        user_data = None
        if order.user_id:
            users_service_url = getenv("USERS_SERVICE_URL", "http://users-service:8001")
            try:
                async with httpx.AsyncClient() as client:
                    user_response = await client.get(
                        f"{users_service_url}/users/{order.user_id}",
                        timeout=5.0
                    )
                    if user_response.status_code == 200:
                        user_data = user_response.json()
                        user_tgid = user_data.get("tgid")
                        user_vkid = user_data.get("vkid")
            except Exception:
                pass
        
        delivery_info = None
        if order.delivery:
            delivery_info = DeliveryInfoForAdminClass(
                delivery_status_id=order.delivery.delivery_status_id,
                delivery_status_name=order.delivery.delivery_status.name if order.delivery.delivery_status else None,
                additional_info=order.delivery.additional_info
            )
        return AdminOrderResponse(
            id=order.id,
            user_id=order.user_id,
            order_data=await admin_order_data_for_response(session, order_id=order.id, order_data=order.order_data),
            status=order.status,
            created_at=order.created_at,
            updated_at=order.updated_at,
            paid_amount=float(order.paid_amount),
            order_total=_order_total_payable(order),
            tracking_number=getattr(order, "tracking_number", None),
            cancel_reason=order.cancel_reason,
            refund_on_cancel=order.refund_on_cancel,
            user_tgid=user_tgid,
            user_vkid=user_vkid,
            delivery=delivery_info,
            phone_number=order.phone_number,
            is_from_stock=order.is_from_stock,
            recipient_name=getattr(order, "recipient_name", None),
        )

    # Запрещаем изменение статуса с "отменен" или "завершен"
    if old_status in ["отменен", "завершен"]:
        raise HTTPException(
            status_code=400,
            detail=f"Нельзя изменить статус заказа с '{old_status}' на другой статус"
        )
    
    # Обрабатываем остатки/резервы при изменении статуса (для заказов со склада)
    if order.is_from_stock and request.new_status == "отменен":
        from api.products.routers.orders import _increase_stock_quantities
        await _increase_stock_quantities(order, session)
        logger.info(f"Резервы сняты для отмененного заказа со склада {order.id}")
    if order.is_from_stock and request.new_status in ("Собран", "завершен"):
        from api.products.routers.orders import _convert_reservations_to_deduction
        await _convert_reservations_to_deduction(order_id, session)
        logger.info(f"Резервы переведены в списание для заказа со склада {order.id}")
    
    if request.new_status == "завершен":
        amount_due = compute_order_amount_due(
            order.order_data,
            float(getattr(order, "tryon_discount_rub", 0) or 0),
            delivery_cost_from_order_snapshot(order.order_data),
            exclude_returned=True,
        )
        paid_amount_decimal = Decimal(str(order.paid_amount))
        amount_due_decimal = Decimal(str(amount_due))
        if paid_amount_decimal < (amount_due_decimal - Decimal("0.01")):
            raise HTTPException(
                status_code=400,
                detail=f"Нельзя перевести заказ в статус 'завершен': не внесена полная оплата. "
                       f"Внесено: {float(order.paid_amount):.2f} ₽, к оплате (товары − примерки + доставка): {amount_due:.2f} ₽"
            )
    
    # Обработка причины отмены
    if request.new_status == "отменен":
        # Для заказов без привязанного пользователя причина отмены не обязательна
        if order.user_id is None:
            order.cancel_reason = request.cancel_reason.strip() if request.cancel_reason and request.cancel_reason.strip() else None
        else:
            if not request.cancel_reason or not request.cancel_reason.strip():
                raise HTTPException(status_code=400, detail="Не указана причина отмены заказа")
            order.cancel_reason = request.cancel_reason.strip()
    else:
        order.cancel_reason = None
    
    # Обработка перехода в статус "в работе" - подтверждаем платежи (capture) для заказов по предзаказу
    # Для заказов по предзаказу платежи подтверждаются при переходе в "в работе"
    # Для заказов из наличия платежи подтверждаются при переходе в "Собран"
    if request.new_status == "в работе" and old_status != "в работе" and not order.is_from_stock:
        logger.info(f"Заказ {order.id} (предзаказ) переходит в статус 'в работе': начинаем подтверждение платежей (capture)")
        try:
            async with httpx.AsyncClient() as client:
                logger.info(f"Вызываем capture-order-payments для заказа {order.id} через {FINANCE_SERVICE_URL}/internal/payments/capture-order-payments")
                capture_response = await client.post(
                    f"{FINANCE_SERVICE_URL}/internal/payments/capture-order-payments",
                    json={"order_id": order.id},
                    headers={"X-Internal-Token": INTERNAL_TOKEN},
                    timeout=30.0
                )
                logger.info(f"Ответ от finance-service для заказа {order.id}: статус {capture_response.status_code}, тело: {capture_response.text}")
                if capture_response.status_code == 200:
                    result = capture_response.json()
                    logger.info(f"Платежи заказа {order.id} подтверждены: {result.get('captured_count', 0)} из {result.get('total_count', 0)}. Ошибки: {result.get('errors', [])}")
                else:
                    logger.warning(f"Ошибка при подтверждении платежей заказа {order.id}: {capture_response.status_code} - {capture_response.text}")
        except Exception as e:
            logger.error(f"Ошибка при подтверждении платежей заказа {order.id}: {e}", exc_info=True)
            # Не прерываем выполнение, но логируем ошибку
    
    # Обработка перехода в статус "Собран" - подтверждаем платежи (capture) для заказов из наличия
    # Для заказов из наличия платежи подтверждаются при переходе в "Собран"
    if request.new_status == "Собран" and old_status != "Собран" and order.is_from_stock:
        logger.info(f"Заказ {order.id} переходит в статус 'Собран': начинаем подтверждение платежей (capture)")
        try:
            async with httpx.AsyncClient() as client:
                logger.info(f"Вызываем capture-order-payments для заказа {order.id} через {FINANCE_SERVICE_URL}/internal/payments/capture-order-payments")
                capture_response = await client.post(
                    f"{FINANCE_SERVICE_URL}/internal/payments/capture-order-payments",
                    json={"order_id": order.id},
                    headers={"X-Internal-Token": INTERNAL_TOKEN},
                    timeout=30.0
                )
                logger.info(f"Ответ от finance-service для заказа {order.id}: статус {capture_response.status_code}, тело: {capture_response.text}")
                if capture_response.status_code == 200:
                    result = capture_response.json()
                    logger.info(f"Платежи заказа {order.id} подтверждены: {result.get('captured_count', 0)} из {result.get('total_count', 0)}. Ошибки: {result.get('errors', [])}")
                else:
                    logger.warning(f"Ошибка при подтверждении платежей заказа {order.id}: {capture_response.status_code} - {capture_response.text}")
        except Exception as e:
            logger.error(f"Ошибка при подтверждении платежей заказа {order.id}: {e}", exc_info=True)
            # Не прерываем выполнение, но логируем ошибку
    
    # Обработка перехода в "отменен": отменяем платежи в холде (cancel) и делаем полный возврат по принятым платежам
    if request.new_status == "отменен":
        try:
            async with httpx.AsyncClient() as client:
                cancel_response = await client.post(
                    f"{FINANCE_SERVICE_URL}/internal/payments/cancel-order-payments",
                    json={"order_id": order.id},
                    headers={"X-Internal-Token": INTERNAL_TOKEN},
                    timeout=30.0
                )
                if cancel_response.status_code == 200:
                    result = cancel_response.json()
                    logger.info(f"Платежи заказа {order.id} отменены: {result.get('canceled_count', 0)} из {result.get('total_count', 0)}")
                else:
                    logger.warning(f"Ошибка при отмене платежей заказа {order.id}: {cancel_response.status_code} - {cancel_response.text}")
        except Exception as e:
            logger.error(f"Ошибка при отмене платежей заказа {order.id}: {e}", exc_info=True)
        # Полный возврат по заказу, если были внесённые средства (принятые платежи)
        if float(order.paid_amount) > 0:
            try:
                async with httpx.AsyncClient() as client:
                    refund_response = await client.post(
                        f"{FINANCE_SERVICE_URL}/internal/refunds/order",
                        json={
                            "order_id": order.id,
                            "reason": (order.cancel_reason or "").strip() or "Отмена заказа",
                        },
                        headers={"X-Internal-Token": INTERNAL_TOKEN},
                        timeout=30.0,
                    )
                    if refund_response.status_code == 200:
                        data = refund_response.json()
                        refunded = data.get("refunded_amount", 0)
                        from decimal import Decimal as Dec
                        order.paid_amount = Dec(str(max(0, float(order.paid_amount) - refunded)))
                        order.refund_on_cancel = True
                        logger.info(f"По заказу {order.id} выполнен возврат при отмене: {refunded} ₽")
                    else:
                        logger.warning(f"Ошибка возврата при отмене заказа {order.id}: {refund_response.status_code} - {refund_response.text}")
            except Exception as e:
                logger.error(f"Ошибка при возврате по заказу {order.id}: {e}", exc_info=True)
    
    # ТОВАРЫ УЖЕ СПИСАНЫ при создании заказа со статусом "в работе", поэтому не списываем при статусе "Собран"
    # Обработка возврата товаров при отмене уже выполнена выше на строке 1818-1821 для всех отмененных заказов со склада
    # (без проверки на статус "Собран" - всегда возвращаем товары при отмене)

    if request.new_status == "завершен" and order.user_id:
        try:
            await complete_tryon_for_order(order.id, order.user_id)
            order.tryon_discount_settled = True
        except Exception as e:
            logger.error("tryon complete для заказа %s: %s", order.id, e, exc_info=True)
            raise HTTPException(
                status_code=502,
                detail="Не удалось зафиксировать примерки в профиле пользователя. Повторите попытку.",
            )
    if request.new_status == "отменен" and order.user_id:
        await release_tryon_for_order(order.id, order.user_id)
    if request.new_status == "отменен":
        await delete_promo_redemptions_for_order(session, order.id)

    order.status = request.new_status
    await session.commit()
    
    # После commit нужно перезагрузить заказ с предзагруженными связанными объектами
    # чтобы избежать проблем с lazy loading при обращении к order.delivery
    result = await session.execute(
        select(Order)
        .options(joinedload(Order.delivery).joinedload(OrderDelivery.delivery_status))
        .where(Order.id == order_id)
    )
    order = result.scalar_one_or_none()
    
    # Уведомляем finance-service об изменении статуса заказа
    # (для обновления балансов при завершении или отмене без возврата)
    if old_status != request.new_status:
        try:
            # Получаем user_id админа из JWT payload (admin уже содержит данные из verify_jwt_token)
            admin_user_id = admin.get("user_id") if isinstance(admin, dict) else None
            
            async with httpx.AsyncClient() as client:
                finance_response = await client.post(
                    f"{FINANCE_SERVICE_URL}/internal/orders/status-update",
                    json={
                        "order_id": order.id,
                        "old_status": old_status,
                        "new_status": request.new_status,
                        "paid_amount": float(order.paid_amount),
                        "refund_on_cancel": order.refund_on_cancel,
                        "admin_user_id": admin_user_id
                    },
                    headers={"X-Internal-Token": INTERNAL_TOKEN},
                    timeout=10.0
                )
                if finance_response.status_code == 200:
                    logger.info(f"Finance service обновлен для заказа #{order.id}")
                else:
                    logger.warning(f"Finance service вернул ошибку для заказа #{order.id}: {finance_response.status_code} - {finance_response.text}")
        except Exception as e:
            # Не прерываем выполнение, если finance-service недоступен
            logger.error(f"Ошибка при уведомлении finance-service об изменении статуса заказа #{order.id}: {e}", exc_info=True)
    
    user_tgid = None
    user_vkid = None
    user_data = None
    if order.user_id:
        try:
            async with httpx.AsyncClient() as client:
                user_response = await client.get(
                    f"{USERS_SERVICE_URL}/users/{order.user_id}",
                    timeout=5.0
                )
                if user_response.status_code == 200:
                    user_data = user_response.json()
                    user_tgid = user_data.get("tgid")
                    user_vkid = user_data.get("vkid")
        except Exception as e:
            logger.warning(f"Не удалось получить информацию о пользователе {order.user_id}: {e}")
    
    # Уведомления всегда в Telegram (миниап только в TG)
    if old_status != request.new_status and user_tgid:
        if request.new_status == "завершен":
            text = (
                f"✅ Ваш заказ #{order.id} завершен!\n\n"
                "Спасибо, что выбрали MatchWear 💛\n\n"
                "Нам будет очень полезен твой отзыв о качестве вещей и работе магазина.\n"
                f"{MINIAPP_REVIEW_HINT}\n"
                f"{REVIEW_CHANNEL_DUPLICATES_HINT}\n"
                "(если ссылка на канал не открывается — скопируй её в поиск Telegram).\n\n"
                f"{SUPPORT_VIA_CHANNEL_HINT}"
            )
            await send_telegram_message(user_tgid, text)
        elif request.new_status == "Собран":
            text = (
                f"📦 Ваш заказ #{order.id} собран и готов.\n\n"
                "Проверьте детали в разделе заказов мини-приложения.\n"
                f"{MINIAPP_ORDERS_HINT}\n\n"
                f"{SUPPORT_VIA_CHANNEL_HINT}"
            )
            await send_telegram_message(user_tgid, text)
        elif request.new_status == "отменен":
            reason_text = order.cancel_reason or "Причина не указана."
            text = (
                f"❌ Ваш заказ #{order.id} был отменен.\n\n"
                f"Причина: {reason_text}\n\n"
                f"{SUPPORT_VIA_CHANNEL_HINT}"
            )
            await send_telegram_message(user_tgid, text)
    
    delivery_info = None
    if order.delivery:
        delivery_info = DeliveryInfoForAdminClass(
            delivery_status_id=order.delivery.delivery_status_id,
            delivery_status_name=order.delivery.delivery_status.name if order.delivery.delivery_status else None,
            additional_info=order.delivery.additional_info
        )
    return AdminOrderResponse(
        id=order.id,
        user_id=order.user_id,
        order_data=await admin_order_data_for_response(session, order_id=order.id, order_data=order.order_data),
        status=order.status,
        created_at=order.created_at,
        updated_at=order.updated_at,
        paid_amount=float(order.paid_amount),
        order_total=_order_total_payable(order),
        tracking_number=getattr(order, "tracking_number", None),
        delivery=delivery_info,
        phone_number=order.phone_number,
        is_from_stock=order.is_from_stock,
        cancel_reason=order.cancel_reason,
        refund_on_cancel=order.refund_on_cancel,
        user_tgid=user_tgid,
        user_vkid=user_vkid
    )


@router.post("/admin/orders/{order_id}/paid-amount", response_model=AdminOrderResponse)
async def update_order_paid_amount(
    order_id: int,
    request: UpdatePaidAmountRequest,
    admin = Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    """Обновить внесенные средства заказа (для админов)"""
    if request.paid_amount < 0:
        raise HTTPException(status_code=400, detail="Внесенные средства не могут быть отрицательными")
    
    result = await session.execute(select(Order).where(Order.id == order_id))
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    
    order.paid_amount = DecimalType(str(request.paid_amount))
    await session.commit()
    await session.refresh(order)
    
    user_tgid = None
    user_vkid = None
    user_data = None
    if order.user_id:
        users_service_url = getenv("USERS_SERVICE_URL", "http://users-service:8001")
        try:
            async with httpx.AsyncClient() as client:
                user_response = await client.get(
                    f"{users_service_url}/users/{order.user_id}",
                    timeout=5.0
                )
                if user_response.status_code == 200:
                    user_data = user_response.json()
                    user_tgid = user_data.get("tgid")
                    user_vkid = user_data.get("vkid")
        except Exception:
            pass
    delivery_info = None
    if order.delivery:
        delivery_info = DeliveryInfoForAdmin(
            delivery_status_id=order.delivery.delivery_status_id,
            delivery_status_name=order.delivery.delivery_status.name if order.delivery.delivery_status else None,
            additional_info=order.delivery.additional_info
        )
    return AdminOrderResponse(
        id=order.id,
        user_id=order.user_id,
        order_data=await admin_order_data_for_response(session, order_id=order.id, order_data=order.order_data),
        status=order.status,
        created_at=order.created_at,
        updated_at=order.updated_at,
        paid_amount=float(order.paid_amount),
        order_total=_order_total_payable(order),
        tracking_number=getattr(order, "tracking_number", None),
        delivery=delivery_info,
        phone_number=order.phone_number,
        is_from_stock=order.is_from_stock,
        cancel_reason=order.cancel_reason,
        refund_on_cancel=order.refund_on_cancel,
        user_tgid=user_tgid,
        user_vkid=user_vkid,
        recipient_name=getattr(order, "recipient_name", None),
    )


@router.get("/admin/orders/{order_id}/payments")
async def get_order_payments(
    order_id: int,
    admin = Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    """Получить все платежи по заказу"""
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(
                f"{FINANCE_SERVICE_URL}/payments/order/{order_id}",
                headers={
                    "X-Internal-Token": INTERNAL_TOKEN
                },
                timeout=10.0
            )
            if response.status_code == 200:
                payment_data = response.json()
                # Если это один платеж, возвращаем список
                if isinstance(payment_data, dict):
                    return [payment_data]
                return payment_data
            elif response.status_code == 404:
                return []
            else:
                logger.warning(f"Ошибка при получении платежей для заказа {order_id}: {response.status_code} - {response.text}")
                return []
    except Exception as e:
        logger.error(f"Ошибка при получении платежей для заказа {order_id}: {e}", exc_info=True)
        return []


@router.get("/admin/orders/{order_id}/payments/{payment_id}/receipt-pdf")
async def admin_download_order_payment_receipt_pdf(
    order_id: int,
    payment_id: int,
    admin=Depends(check_orders_access),
):
    """
    PDF-справка по позициям чека, сохранённым при создании платежа (прокси в finance-service).
    """
    try:
        async with httpx.AsyncClient() as client:
            r = await client.get(
                f"{FINANCE_SERVICE_URL}/internal/payments/{payment_id}/receipt-pdf",
                params={"order_id": order_id},
                headers={"X-Internal-Token": INTERNAL_TOKEN},
                timeout=60.0,
            )
        if r.status_code == 404:
            try:
                detail = r.json().get("detail", "Не найдено")
            except Exception:
                detail = "Не найдено"
            raise HTTPException(status_code=404, detail=detail)
        if r.status_code != 200:
            logger.warning(
                "receipt-pdf finance HTTP %s: %s",
                r.status_code,
                (r.text or "")[:500],
            )
            raise HTTPException(
                status_code=502,
                detail="Не удалось получить PDF из сервиса финансов",
            )
        headers = {}
        cd = r.headers.get("content-disposition")
        if cd:
            headers["Content-Disposition"] = cd
        return Response(
            content=r.content,
            media_type=r.headers.get("content-type", "application/pdf"),
            headers=headers,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка receipt-pdf заказ {order_id} платёж {payment_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Ошибка при получении PDF чека")


class UpdateOrderDeliveryDataRequest(BaseModel):
    phone_number: Optional[str] = None
    tracking_number: Optional[str] = None


class UpdateOrderDeliverySnapshotRequest(BaseModel):
    """Обновить данные доставки, выбранные пользователем (адрес, тип, стоимость и т.д.)."""
    recipient_name: Optional[str] = None
    address: Optional[str] = None
    postal_code: Optional[str] = None
    city_code: Optional[int] = None
    delivery_cost_rub: Optional[float] = None
    delivery_method_code: Optional[str] = None
    cdek_delivery_point_code: Optional[str] = None


_CARRIER_SETTLEMENT_MONTH_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")
_ALLOWED_CARRIER_SETTLEMENT = frozenset({"CDEK", "LOCAL_COURIER", "OTHER"})


class PatchDeliveryCarrierSettlementRequest(BaseModel):
    """
    Учёт фактической суммы к оплате перевозчику (СДЭК / локальный курьер / прочее) — отдельно от суммы в чеке клиента.
    Используется в финансовом дашборде: сводка по месяцам.
    """
    clear: bool = False
    payable_rub: Optional[float] = None
    carrier: Optional[str] = None  # CDEK | LOCAL_COURIER | OTHER
    accrual_month: Optional[str] = None  # YYYY-MM
    note: Optional[str] = None


async def patch_order_delivery_carrier_settlement(
    order_id: int,
    request: PatchDeliveryCarrierSettlementRequest,
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session),
):
    raise HTTPException(
        status_code=410,
        detail="Учёт оплат перевозчику отключен в StoreKPLite.",
    )


@router.patch("/admin/orders/{order_id}/delivery-snapshot", response_model=AdminOrderResponse)
async def update_order_delivery_snapshot(
    order_id: int,
    request: UpdateOrderDeliverySnapshotRequest,
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session),
):
    """Обновить снимок доставки заказа (адрес, тип доставки, стоимость) и ФИО получателя."""
    result = await session.execute(
        select(Order).options(joinedload(Order.delivery).joinedload(OrderDelivery.delivery_status)).where(Order.id == order_id)
    )
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    data = order.order_data or {}
    snapshot = dict(data.get("delivery_snapshot") or {})
    if request.address is not None:
        snapshot["address"] = request.address.strip() if request.address else None
    if request.postal_code is not None:
        snapshot["postal_code"] = request.postal_code.strip() if request.postal_code else None
    if request.city_code is not None:
        snapshot["city_code"] = request.city_code
    if request.delivery_cost_rub is not None:
        snapshot["delivery_cost_rub"] = request.delivery_cost_rub
    if request.delivery_method_code is not None:
        snapshot["delivery_method_code"] = request.delivery_method_code.strip() if request.delivery_method_code else None
    if request.cdek_delivery_point_code is not None:
        v = request.cdek_delivery_point_code.strip()
        snapshot["cdek_delivery_point_code"] = v or None
    data["delivery_snapshot"] = snapshot
    order.order_data = data
    flag_modified(order, "order_data")
    if request.recipient_name is not None:
        order.recipient_name = request.recipient_name.strip() if request.recipient_name else None
    await session.commit()
    await session.refresh(order)
    return await _build_admin_order_response(session, order)


async def _build_admin_order_response(session: AsyncSession, order) -> AdminOrderResponse:
    """Построить AdminOrderResponse для заказа (с user_tgid/user_vkid, контактами и delivery_info)."""
    display_tgid = None
    display_vkid = None
    user_firstname = None
    user_username = None
    user_email = None
    user_phone = None
    if order.user_id:
        try:
            async with httpx.AsyncClient() as client:
                r = await client.get(f"{USERS_SERVICE_URL}/users/{order.user_id}", timeout=5.0)
                if r.status_code == 200:
                    ud = r.json()
                    display_tgid = ud.get("tgid")
                    display_vkid = ud.get("vkid")
                    user_firstname = ud.get("firstname")
                    user_username = ud.get("username")
                    user_email = ud.get("email")
                    cc = (ud.get("country_code") or "").strip()
                    pl = (ud.get("phone_local") or "").strip()
                    user_phone = (cc + " " + pl).strip() if (cc or pl) else None
        except Exception:
            pass
    delivery_info = None
    if order.delivery:
        delivery_info = DeliveryInfoForAdmin(
            delivery_status_id=order.delivery.delivery_status_id,
            delivery_status_name=order.delivery.delivery_status.name if order.delivery.delivery_status else None,
            additional_info=order.delivery.additional_info,
        )
    return AdminOrderResponse(
        id=order.id,
        user_id=order.user_id,
        order_data=await admin_order_data_for_response(session, order_id=order.id, order_data=order.order_data),
        status=order.status,
        created_at=order.created_at,
        updated_at=order.updated_at,
        paid_amount=float(order.paid_amount),
        order_total=_order_total_payable(order, exclude_returned=True),
        tracking_number=getattr(order, "tracking_number", None),
        delivery=delivery_info,
        phone_number=order.phone_number,
        is_from_stock=order.is_from_stock,
        cancel_reason=order.cancel_reason,
        refund_on_cancel=order.refund_on_cancel,
        user_tgid=display_tgid,
        user_vkid=display_vkid,
        recipient_name=getattr(order, "recipient_name", None),
        user_firstname=user_firstname,
        user_username=user_username,
        user_email=user_email,
        user_phone=user_phone,
    )


@router.patch("/admin/orders/{order_id}/delivery-data", response_model=AdminOrderResponse)
async def update_order_delivery_data(
    order_id: int,
    request: UpdateOrderDeliveryDataRequest,
    admin = Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    """Обновить данные доставки заказа (телефон, трек-номер)"""
    result = await session.execute(
        select(Order).options(joinedload(Order.delivery).joinedload(OrderDelivery.delivery_status)).where(Order.id == order_id)
    )
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    
    if request.phone_number is not None:
        order.phone_number = request.phone_number if request.phone_number else None
    if request.tracking_number is not None:
        order.tracking_number = request.tracking_number if request.tracking_number else None
    
    await session.commit()
    await session.refresh(order)
    
    display_tgid = None
    display_vkid = None
    if order.user_id:
        try:
            async with httpx.AsyncClient() as client:
                user_response = await client.get(
                    f"{USERS_SERVICE_URL}/users/{order.user_id}",
                    timeout=5.0
                )
                if user_response.status_code == 200:
                    user_data = user_response.json()
                    display_tgid = user_data.get("tgid")
                    display_vkid = user_data.get("vkid")
        except Exception as e:
            logger.warning(f"Не удалось получить информацию о пользователе {order.user_id}: {e}")
    
    delivery_info = None
    delivery_result = await session.execute(
        select(OrderDelivery).where(OrderDelivery.order_id == order_id)
    )
    delivery = delivery_result.scalar_one_or_none()
    
    if delivery:
        delivery_status_name = None
        if delivery.delivery_status_id:
            status_result = await session.execute(
                select(DeliveryStatus).where(DeliveryStatus.id == delivery.delivery_status_id)
            )
            status = status_result.scalar_one_or_none()
            if status:
                delivery_status_name = status.name
        
        delivery_info = DeliveryInfoForAdmin(
            delivery_status_id=delivery.delivery_status_id,
            delivery_status_name=delivery_status_name,
            additional_info=delivery.additional_info
        )
    
    return AdminOrderResponse(
        id=order.id,
        user_id=order.user_id,
        order_data=await admin_order_data_for_response(session, order_id=order.id, order_data=order.order_data),
        status=order.status,
        created_at=order.created_at,
        updated_at=order.updated_at,
        paid_amount=float(order.paid_amount),
        order_total=_order_total_payable(order),
        tracking_number=getattr(order, "tracking_number", None),
        delivery=delivery_info,
        phone_number=order.phone_number,
        is_from_stock=order.is_from_stock,
        cancel_reason=order.cancel_reason,
        refund_on_cancel=order.refund_on_cancel,
        user_tgid=display_tgid,
        user_vkid=display_vkid,
        recipient_name=getattr(order, "recipient_name", None),
    )


@router.post("/admin/orders/{order_id}/send-payment-reminder")
async def send_payment_reminder(
    order_id: int,
    admin = Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    """Отправить уведомление пользователю о необходимости доплаты"""
    result = await session.execute(select(Order).where(Order.id == order_id))
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    
    order_total = _order_total_payable(order)
    if order_total is None or order_total <= 0:
        raise HTTPException(status_code=400, detail="У заказа не указана стоимость (нет позиций)")
    
    remaining = float(order_total) - float(order.paid_amount)
    if remaining <= 0:
        raise HTTPException(status_code=400, detail="Заказ полностью оплачен")
    
    # Получаем информацию о пользователе
    user_tgid = None
    user_vkid = None
    if order.user_id:
        try:
            async with httpx.AsyncClient() as client:
                user_response = await client.get(
                    f"{USERS_SERVICE_URL}/users/{order.user_id}",
                    timeout=5.0
                )
                if user_response.status_code == 200:
                    user_data = user_response.json()
                    user_tgid = user_data.get("tgid")
                    user_vkid = user_data.get("vkid")
        except Exception as e:
            logger.warning(f"Не удалось получить информацию о пользователе {order.user_id}: {e}")
    
    if not user_tgid:
        raise HTTPException(status_code=400, detail="У пользователя нет Telegram (уведомления только в TG)")
    
    text = (
        f"📦 Напоминание: по заказу #{order_id} ожидается оплата.\n\n"
        f"💵 Сумма заказа: {order_total:.2f} ₽\n"
        f"💰 Уже внесено: {float(order.paid_amount):.2f} ₽\n"
        f"📝 К оплате: {remaining:.2f} ₽\n\n"
        f"{MINIAPP_ORDERS_HINT}"
    )
    await send_telegram_message(user_tgid, text)
    
    return {"status": "ok", "message": "Уведомление отправлено"}


class CreateManualOrderResponse(AdminOrderResponse):
    """Ответ на создание ручного заказа: deep-link в мини-приложение (заказ уже привязан к пользователю)."""
    order_link: str


async def _manual_order_build_delivery_snapshot(
    delivery: ManualOrderDeliveryInput,
    order_items_for_parcel: List[Dict[str, Any]],
    session: AsyncSession,
) -> Tuple[Dict[str, Any], Optional[str]]:
    """
    Формирует delivery_snapshot и возвращает (snapshot, recipient_name).
    Стоимость для CDEK при отсутствии delivery_cost_rub считается через delivery-service (как при чекауте).
    """
    code = (delivery.delivery_method_code or "").strip().upper()
    if code in ("LOCAL_COURIER",):
        code = "COURIER_LOCAL"
    if code in ("LOCAL_PICKUP_POINT",):
        code = "PICKUP_LOCAL"

    recipient = (delivery.recipient_name or "").strip() or None
    snap: Dict[str, Any] = {"delivery_method_code": code}

    if code == "PICKUP_LOCAL":
        if not delivery.local_pickup_point_id:
            raise HTTPException(status_code=400, detail="Для локального ПВЗ укажите local_pickup_point_id")
        try:
            async with httpx.AsyncClient(timeout=10.0) as client:
                r = await client.get(f"{DELIVERY_SERVICE_URL.rstrip('/')}/local-pickup-points")
        except httpx.RequestError as e:
            logger.warning("manual order: local pickup list: %s", e)
            raise HTTPException(status_code=502, detail="Сервис доставки недоступен")
        if r.status_code != 200:
            raise HTTPException(status_code=502, detail="Не удалось получить список локальных ПВЗ")
        pts = r.json() if isinstance(r.json(), list) else []
        pt = next((p for p in pts if isinstance(p, dict) and int(p.get("id") or 0) == int(delivery.local_pickup_point_id)), None)
        if not pt:
            raise HTTPException(status_code=400, detail="Локальный ПВЗ не найден")
        city = (pt.get("city") or "").strip()
        addr = (pt.get("address") or "").strip()
        snap["address"] = f"{city}, {addr}".strip(", ").strip() or addr
        cost = delivery.delivery_cost_rub
        if cost is None:
            cost = 0.0
        snap["delivery_cost_rub"] = float(cost)
        return snap, recipient

    if code == "COURIER_LOCAL":
        addr = (delivery.address or "").strip()
        if not addr:
            raise HTTPException(status_code=400, detail="Для курьера укажите адрес")
        snap["address"] = addr
        if delivery.postal_code:
            snap["postal_code"] = (delivery.postal_code or "").strip() or None
        cost = delivery.delivery_cost_rub
        if cost is None:
            try:
                async with httpx.AsyncClient(timeout=10.0) as client:
                    calc_r = await client.post(
                        f"{DELIVERY_SERVICE_URL.rstrip('/')}/calculate-cost",
                        json={
                            "parcel": {
                                "weight_gram": 1000,
                                "length_cm": 40,
                                "width_cm": 30,
                                "height_cm": 10,
                            },
                            "delivery_method_code": "COURIER_LOCAL",
                        },
                    )
                if calc_r.status_code == 200:
                    cj = calc_r.json()
                    if cj.get("delivery_cost_rub") is not None:
                        cost = float(cj["delivery_cost_rub"])
            except Exception as e:
                logger.warning("manual order courier cost: %s", e)
        snap["delivery_cost_rub"] = float(cost) if cost is not None else 0.0
        return snap, recipient

    if code == "CDEK_MANUAL":
        addr = (delivery.address or "").strip()
        if not addr:
            raise HTTPException(status_code=400, detail="Для СДЭК (адрес вручную) укажите адрес одной строкой")
        snap["address"] = addr
        if delivery.postal_code:
            snap["postal_code"] = (delivery.postal_code or "").strip() or None
        snap["delivery_cost_rub"] = delivery.delivery_cost_rub
        return snap, recipient

    if code == "CDEK":
        dp = (delivery.cdek_delivery_point_code or "").strip()
        if not dp:
            raise HTTPException(status_code=400, detail="Для СДЭК укажите код ПВЗ (cdek_delivery_point_code)")
        city_code = delivery.delivery_city_code
        if city_code is None:
            cname = (delivery.delivery_city or "").strip()
            if not cname:
                raise HTTPException(
                    status_code=400,
                    detail="Для СДЭК укажите delivery_city_code или название города в delivery_city",
                )
            try:
                async with httpx.AsyncClient(timeout=15.0) as client:
                    pr = await client.get(
                        f"{DELIVERY_SERVICE_URL.rstrip('/')}/pickup-points",
                        params={"city": cname, "country_code": "RU", "limit": 1},
                    )
                if pr.status_code != 200:
                    raise HTTPException(status_code=400, detail="Не удалось определить город для СДЭК")
                plist = pr.json() if isinstance(pr.json(), list) else []
                if not plist or not isinstance(plist[0], dict):
                    raise HTTPException(status_code=400, detail="Город не найден в справочнике СДЭК")
                city_code = plist[0].get("city_code")
            except HTTPException:
                raise
            except Exception as e:
                logger.warning("manual order CDEK city: %s", e)
                raise HTTPException(status_code=502, detail="Ошибка запроса к сервису доставки")
        if city_code is None:
            raise HTTPException(status_code=400, detail="Не удалось получить код города СДЭК")
        try:
            city_code = int(city_code)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Некорректный код города СДЭК")

        snap["city_code"] = city_code
        snap["cdek_delivery_point_code"] = dp
        addr = (delivery.address or "").strip()
        if addr:
            snap["address"] = addr
        if delivery.postal_code:
            snap["postal_code"] = (delivery.postal_code or "").strip() or None

        item_ids = [it.get("item_id") for it in order_items_for_parcel if it.get("item_id")]
        items_by_id: Dict[int, Any] = {}
        if item_ids:
            items_result = await session.execute(select(Item).where(Item.id.in_(item_ids)))
            for it in items_result.scalars().all():
                items_by_id[it.id] = it
        line_items = build_line_items_for_parcel(order_items_for_parcel, items_by_id)
        parcel = aggregate_parcel_dimensions(line_items)

        delivery_cost = delivery.delivery_cost_rub
        cdek_tariff: Optional[int] = None
        if delivery_cost is None:
            try:
                calc_body: Dict[str, Any] = {
                    "parcel": {
                        "weight_gram": parcel["weight_gram"],
                        "length_cm": parcel["length_cm"],
                        "width_cm": parcel["width_cm"],
                        "height_cm": parcel["height_cm"],
                    },
                    "delivery_method_code": "CDEK",
                    "to_city_code": city_code,
                    "cdek_delivery_point_code": dp,
                }
                calc_body.update(cdek_delivery_calc_insurance_extras("CDEK", order_items_for_parcel))
                async with httpx.AsyncClient(timeout=12.0) as client:
                    calc_r = await client.post(
                        f"{DELIVERY_SERVICE_URL.rstrip('/')}/calculate-cost",
                        json=calc_body,
                    )
                if calc_r.status_code == 200:
                    cj = calc_r.json()
                    if cj.get("delivery_cost_rub") is not None:
                        delivery_cost = float(cj["delivery_cost_rub"])
                    tc = cj.get("cdek_tariff_code")
                    if tc is not None:
                        try:
                            cdek_tariff = int(tc)
                        except (TypeError, ValueError):
                            pass
            except Exception as e:
                logger.warning("manual order CDEK calc: %s", e)
        if delivery_cost is not None:
            snap["delivery_cost_rub"] = float(delivery_cost)
        if cdek_tariff is not None:
            snap["cdek_tariff_code"] = cdek_tariff
        return snap, recipient

    raise HTTPException(
        status_code=400,
        detail=f"Неизвестный способ доставки: {delivery.delivery_method_code!r}. Ожидаются PICKUP_LOCAL, COURIER_LOCAL, CDEK, CDEK_MANUAL",
    )


@router.post("/admin/orders/manual/upload-custom-photo")
async def upload_manual_custom_item_photo(
    photo: UploadFile = File(...),
    admin=Depends(check_orders_access),
):
    """Одно фото для кастомной позиции ручного заказа; вернуть относительный путь для поля photo в составе."""
    saved = await save_compressed_image(photo)
    rel = f"uploads/items/{saved.name}"
    return {"file_path": rel}


@router.post("/admin/orders/manual/preview-delivery-cost", response_model=ManualDeliveryPreviewResponse)
async def preview_manual_delivery_cost(
    request: ManualDeliveryPreviewRequest,
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session),
):
    """
    Предпросчёт доставки СДЭК до создания ручного заказа: тот же POST …/calculate-cost,
    что при оформлении в миниаппе (в т.ч. страховая база через cdek_delivery_calc_insurance_extras).
    """
    code = (request.delivery_method_code or "CDEK").strip() or "CDEK"
    if code != "CDEK":
        raise HTTPException(status_code=400, detail="Предпросчёт поддержан только для CDEK")
    dp = (request.cdek_delivery_point_code or "").strip()
    if not dp:
        raise HTTPException(status_code=400, detail="Укажите код ПВЗ СДЭК (cdek_delivery_point_code)")
    try:
        city_code = int(request.delivery_city_code)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Некорректный delivery_city_code")

    preview_rows: List[Dict[str, Any]] = []
    catalog_ids: List[int] = []
    for it in request.items:
        qty = max(0, int(it.quantity or 1))
        if qty == 0:
            continue
        row: Dict[str, Any] = {
            "quantity": qty,
            "price": float(it.price or 0),
            "name": (it.name or "").strip() or "Товар",
        }
        if it.item_id is not None:
            row["item_id"] = int(it.item_id)
            catalog_ids.append(int(it.item_id))
        else:
            if it.estimated_weight_kg is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Для кастомной позиции «{row['name']}» укажите estimated_weight_kg (кг)",
                )
            if it.length_cm is None or it.width_cm is None or it.height_cm is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Для кастомной позиции «{row['name']}» укажите length_cm, width_cm, height_cm",
                )
            try:
                w = float(it.estimated_weight_kg)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail=f"Некорректный вес для «{row['name']}»")
            if w <= 0:
                raise HTTPException(status_code=400, detail=f"Вес должен быть > 0 для «{row['name']}»")
            try:
                lc, wc, hc = int(it.length_cm), int(it.width_cm), int(it.height_cm)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail=f"Некорректные габариты для «{row['name']}»")
            if min(lc, wc, hc) < 1:
                raise HTTPException(status_code=400, detail=f"Габариты должны быть ≥ 1 см для «{row['name']}»")
            row["estimated_weight_kg"] = w
            row["length_cm"] = lc
            row["width_cm"] = wc
            row["height_cm"] = hc
        preview_rows.append(row)

    if not preview_rows:
        raise HTTPException(status_code=400, detail="Добавьте хотя бы одну позицию с ненулевым количеством")

    uniq_ids = list({i for i in catalog_ids})
    items_by_id: Dict[int, Item] = {}
    if uniq_ids:
        items_result = await session.execute(select(Item).where(Item.id.in_(uniq_ids)))
        for obj in items_result.scalars().all():
            items_by_id[obj.id] = obj
        missing = [i for i in uniq_ids if i not in items_by_id]
        if missing:
            raise HTTPException(status_code=400, detail=f"Товары каталога не найдены: {missing[:10]}")

    line_items = build_line_items_for_parcel(preview_rows, items_by_id)
    parcel = aggregate_parcel_dimensions(line_items)
    calc_body: Dict[str, Any] = {
        "parcel": {
            "weight_gram": parcel["weight_gram"],
            "length_cm": parcel["length_cm"],
            "width_cm": parcel["width_cm"],
            "height_cm": parcel["height_cm"],
        },
        "delivery_method_code": "CDEK",
        "to_city_code": city_code,
        "cdek_delivery_point_code": dp,
    }
    calc_body.update(cdek_delivery_calc_insurance_extras("CDEK", preview_rows))
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            calc_r = await client.post(
                f"{DELIVERY_SERVICE_URL.rstrip('/')}/calculate-cost",
                json=calc_body,
            )
    except Exception as e:
        logger.warning("manual preview CDEK calc: %s", e)
        raise HTTPException(status_code=502, detail="Сервис доставки недоступен или не ответил")
    if calc_r.status_code != 200:
        raise HTTPException(status_code=502, detail="Не удалось рассчитать доставку")
    cj = calc_r.json() if calc_r.content else {}
    if not isinstance(cj, dict):
        cj = {}
    dc = cj.get("delivery_cost_rub")
    tc = cj.get("cdek_tariff_code")
    cdek_tariff_code: Optional[int] = None
    if tc is not None:
        try:
            cdek_tariff_code = int(tc)
        except (TypeError, ValueError):
            cdek_tariff_code = None
    base = cj.get("cdek_delivery_sum_base_rub")
    total = cj.get("cdek_total_sum_rub")
    return ManualDeliveryPreviewResponse(
        delivery_cost_rub=float(dc) if dc is not None else None,
        cdek_tariff_code=cdek_tariff_code,
        cdek_delivery_sum_base_rub=float(base) if base is not None else None,
        cdek_total_sum_rub=float(total) if total is not None else None,
    )


@router.post("/admin/orders/manual", response_model=CreateManualOrderResponse)
async def create_manual_order(
    request: CreateManualOrderRequest,
    admin = Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    """Создать заказ вручную: пользователь обязателен, статус «Ожидает», предзаказ (как под заказ, не со склада)."""
    if not request.items:
        raise HTTPException(status_code=400, detail="Заказ должен содержать хотя бы один товар")

    try:
        async with httpx.AsyncClient() as client:
            user_resp = await client.get(
                f"{USERS_SERVICE_URL.rstrip('/')}/users/{request.user_id}",
                timeout=8.0,
            )
    except httpx.RequestError as e:
        logger.warning("Не удалось проверить пользователя %s: %s", request.user_id, e)
        raise HTTPException(status_code=502, detail="Сервис пользователей недоступен")
    if user_resp.status_code == 404:
        raise HTTPException(status_code=400, detail="Пользователь не найден")
    if user_resp.status_code != 200:
        raise HTTPException(status_code=502, detail="Не удалось проверить пользователя")

    user_tgid = None
    user_vkid = None
    try:
        uj = user_resp.json()
        if isinstance(uj, dict):
            user_tgid = uj.get("tgid")
            user_vkid = uj.get("vkid")
    except Exception:
        pass

    order_items: List[Dict[str, Any]] = []
    total_price = Decimal("0")

    for item in request.items:
        item_data: Dict[str, Any] = {
            "name": item.name,
            "quantity": item.quantity,
            "price": item.price,
            "stock_type": "preorder",
        }

        if item.item_id:
            item_data["item_id"] = item.item_id
            item_result = await session.execute(
                select(Item).where(Item.id == item.item_id)
            )
            catalog_item = item_result.scalar_one_or_none()
            if catalog_item and catalog_item.link:
                item_data["link"] = catalog_item.link

        if item.size:
            item_data["size"] = item.size

        if item.link and "link" not in item_data:
            item_data["link"] = item.link

        if item.photo:
            if item.item_id:
                raise HTTPException(
                    status_code=400,
                    detail="Поле photo допускается только для кастомных позиций без каталога (без item_id)",
                )
            if not _is_safe_manual_custom_photo_relative_path(item.photo):
                raise HTTPException(status_code=400, detail="Некорректный путь к фото кастомного товара")
            disk_path = UPLOAD_DIR / Path(item.photo.replace("\\", "/")).name
            if not disk_path.is_file():
                raise HTTPException(status_code=400, detail="Файл фото не найден — загрузите фото заново")
            item_data["photo"] = item.photo.strip().replace("\\", "/")

        if not item.item_id:
            if item.estimated_weight_kg is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Для кастомной позиции «{item.name}» укажите вес (estimated_weight_kg), кг",
                )
            if item.length_cm is None or item.width_cm is None or item.height_cm is None:
                raise HTTPException(
                    status_code=400,
                    detail=f"Для кастомной позиции «{item.name}» укажите габариты length_cm, width_cm, height_cm",
                )
            try:
                w_kg = float(item.estimated_weight_kg)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail=f"Некорректный вес для «{item.name}»")
            if w_kg <= 0:
                raise HTTPException(status_code=400, detail=f"Вес должен быть > 0 для «{item.name}»")
            try:
                lc, wc, hc = int(item.length_cm), int(item.width_cm), int(item.height_cm)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail=f"Некорректные габариты для «{item.name}»")
            if min(lc, wc, hc) < 1:
                raise HTTPException(status_code=400, detail=f"Габариты должны быть целыми числами ≥ 1 см для «{item.name}»")
            item_data["estimated_weight_kg"] = w_kg
            item_data["length_cm"] = lc
            item_data["width_cm"] = wc
            item_data["height_cm"] = hc

        if not item.item_id:
            cn = normalize_optional_name(item.chinese_name)
            if cn:
                item_data["chinese_name"] = cn
        elif normalize_optional_name(item.chinese_name):
            raise HTTPException(
                status_code=400,
                detail="Поле chinese_name допускается только для кастомных позиций (без item_id)",
            )

        order_items.append(item_data)
        total_price += Decimal(str(item.quantity)) * Decimal(str(item.price))

    order_data: Dict[str, Any] = {"items": order_items}
    recipient_for_order: Optional[str] = None
    if request.delivery:
        snap_del, recipient_for_order = await _manual_order_build_delivery_snapshot(
            request.delivery, order_items, session
        )
        order_data["delivery_snapshot"] = snap_del

    paid_amount = Decimal("0")
    if request.is_paid:
        paid_amount = total_price

    new_order = Order(
        user_id=request.user_id,
        order_data=order_data,
        status="Ожидает",
        phone_number=request.phone_number,
        paid_amount=paid_amount,
        is_from_stock=False,
        recipient_name=recipient_for_order,
    )

    session.add(new_order)
    await session.commit()
    await session.refresh(new_order)

    bot_username = TELEGRAM_BOT_USERNAME or "bot"
    order_link = f"https://t.me/{bot_username}?startapp=order_{new_order.id}"

    return CreateManualOrderResponse(
        id=new_order.id,
        user_id=new_order.user_id,
        order_data=await admin_order_data_for_response(
            session, order_id=new_order.id, order_data=new_order.order_data
        ),
        status=new_order.status,
        created_at=new_order.created_at,
        updated_at=new_order.updated_at,
        paid_amount=float(new_order.paid_amount),
        order_total=compute_order_total(new_order.order_data),
        tracking_number=getattr(new_order, "tracking_number", None),
        delivery=None,
        phone_number=new_order.phone_number,
        is_from_stock=new_order.is_from_stock,
        cancel_reason=new_order.cancel_reason,
        refund_on_cancel=new_order.refund_on_cancel,
        user_tgid=user_tgid,
        user_vkid=user_vkid,
        order_link=order_link,
    )


class UpdateOrderItemsRequest(BaseModel):
    """Удалить позиции из заказа по индексам (0-based)."""
    remove_indices: List[int]  # индексы в order_data["items"], которые нужно убрать


@router.patch("/admin/orders/{order_id}/items", response_model=AdminOrderResponse)
async def update_order_items(
    order_id: int,
    request: UpdateOrderItemsRequest,
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    """Удалить указанные позиции из заказа. Пересчитывается итог. Если paid_amount > новый итог — можно выполнить рефаунд через POST /refund."""
    result = await session.execute(
        select(Order).options(joinedload(Order.delivery).joinedload(OrderDelivery.delivery_status)).where(Order.id == order_id)
    )
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    if order.status in ("отменен", "завершен"):
        raise HTTPException(status_code=400, detail="Нельзя редактировать состав отменённого или завершённого заказа")
    data = order.order_data or {}
    items = list(data.get("items") or [])
    if not items:
        raise HTTPException(status_code=400, detail="В заказе нет позиций")
    to_remove = set(request.remove_indices)
    new_items = [item for i, item in enumerate(items) if i not in to_remove]
    if not new_items:
        raise HTTPException(status_code=400, detail="Нельзя удалить все позиции")
    data = {**data, "items": new_items}
    order.order_data = data
    flag_modified(order, "order_data")
    await session.commit()
    await session.refresh(order)
    order_total = _order_total_payable(order)
    user_tgid = None
    user_vkid = None
    if order.user_id:
        try:
            async with httpx.AsyncClient() as client:
                r = await client.get(
                    f"{getenv('USERS_SERVICE_URL', 'http://users-service:8001')}/users/{order.user_id}",
                    timeout=5.0
                )
                if r.status_code == 200:
                    ud = r.json()
                    user_tgid = ud.get("tgid")
                    user_vkid = ud.get("vkid")
        except Exception:
            pass
    delivery_info = None
    if order.delivery:
        delivery_info = DeliveryInfoForAdminClass(
            delivery_status_id=order.delivery.delivery_status_id,
            delivery_status_name=order.delivery.delivery_status.name if order.delivery.delivery_status else None,
            additional_info=order.delivery.additional_info
        )
    response = AdminOrderResponse(
        id=order.id,
        user_id=order.user_id,
        order_data=await admin_order_data_for_response(session, order_id=order.id, order_data=order.order_data),
        status=order.status,
        created_at=order.created_at,
        updated_at=order.updated_at,
        paid_amount=float(order.paid_amount),
        order_total=order_total,
        tracking_number=getattr(order, "tracking_number", None),
        delivery=delivery_info,
        phone_number=order.phone_number,
        is_from_stock=order.is_from_stock,
        cancel_reason=order.cancel_reason,
        refund_on_cancel=order.refund_on_cancel,
        user_tgid=user_tgid,
        user_vkid=user_vkid,
        recipient_name=getattr(order, "recipient_name", None),
    )
    return response


class MarkOrderItemsReturnedRequest(BaseModel):
    """Пометить позиции заказа как возвращённые (не показываются в составе на фронте, не входят в итог)."""
    returned_indices: List[int]  # индексы в order_data["items"] (0-based)


@router.patch("/admin/orders/{order_id}/items/returned", response_model=AdminOrderResponse)
async def mark_order_items_returned(
    order_id: int,
    request: MarkOrderItemsReturnedRequest,
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    """Пометить указанные позиции заказа как возвращённые. На фронте они не показываются в составе; итог пересчитывается без них."""
    result = await session.execute(
        select(Order).options(joinedload(Order.delivery).joinedload(OrderDelivery.delivery_status)).where(Order.id == order_id)
    )
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    data = order.order_data or {}
    items = list(data.get("items") or [])
    if not items:
        raise HTTPException(status_code=400, detail="В заказе нет позиций")
    to_mark = set(request.returned_indices)
    for i in to_mark:
        if 0 <= i < len(items):
            items[i] = {**items[i], "returned": True}
    data = {**data, "items": items}
    order.order_data = data
    flag_modified(order, "order_data")
    await session.commit()
    await session.refresh(order)
    order_total = _order_total_payable(order, exclude_returned=True)
    user_tgid = None
    user_vkid = None
    if order.user_id:
        try:
            async with httpx.AsyncClient() as client:
                r = await client.get(
                    f"{getenv('USERS_SERVICE_URL', 'http://users-service:8001')}/users/{order.user_id}",
                    timeout=5.0
                )
                if r.status_code == 200:
                    ud = r.json()
                    user_tgid = ud.get("tgid")
                    user_vkid = ud.get("vkid")
        except Exception:
            pass
    delivery_info = None
    if order.delivery:
        delivery_info = DeliveryInfoForAdminClass(
            delivery_status_id=order.delivery.delivery_status_id,
            delivery_status_name=order.delivery.delivery_status.name if order.delivery.delivery_status else None,
            additional_info=order.delivery.additional_info
        )
    return AdminOrderResponse(
        id=order.id,
        user_id=order.user_id,
        order_data=await admin_order_data_for_response(session, order_id=order.id, order_data=order.order_data),
        status=order.status,
        created_at=order.created_at,
        updated_at=order.updated_at,
        paid_amount=float(order.paid_amount),
        order_total=order_total,
        tracking_number=getattr(order, "tracking_number", None),
        delivery=delivery_info,
        phone_number=order.phone_number,
        is_from_stock=order.is_from_stock,
        cancel_reason=order.cancel_reason,
        refund_on_cancel=order.refund_on_cancel,
        user_tgid=user_tgid,
        user_vkid=user_vkid,
        recipient_name=getattr(order, "recipient_name", None),
    )


class RefundOrderRequest(BaseModel):
    amount: Optional[float] = None  # сумма возврата; если не указана — вся внесённая по succeeded платежам
    reason: Optional[str] = None   # причина: вещь возвращена и т.д.


@router.post("/admin/orders/{order_id}/refund", response_model=AdminOrderResponse)
async def refund_order(
    order_id: int,
    request: RefundOrderRequest,
    admin=Depends(check_orders_access),
    session: AsyncSession = Depends(get_session)
):
    """Выполнить возврат по заказу (частичный или полный). Работает для платежей уже не в холде и для завершённых заказов. ЮKassa возврат + лог."""
    result = await session.execute(
        select(Order).options(joinedload(Order.delivery).joinedload(OrderDelivery.delivery_status)).where(Order.id == order_id)
    )
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Заказ не найден")
    amount = None
    if request.amount is not None and request.amount > 0:
        amount = request.amount
    reason = (request.reason or "").strip() or "Возврат по заказу (вещь возвращена)"
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(
                f"{FINANCE_SERVICE_URL}/internal/refunds/order",
                json={"order_id": order_id, "amount": amount, "reason": reason},
                headers={"X-Internal-Token": INTERNAL_TOKEN}
            )
            if resp.status_code != 200:
                err = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {}
                raise HTTPException(
                    status_code=resp.status_code,
                    detail=err.get("detail", resp.text)
                )
            data = resp.json()
    except httpx.HTTPStatusError as e:
        if hasattr(e, "response") and e.response is not None:
            try:
                err = e.response.json()
                raise HTTPException(status_code=e.response.status_code, detail=err.get("detail", str(e)))
            except Exception:
                raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
        raise
    refunded = data.get("refunded_amount", 0)
    from decimal import Decimal as Dec
    order.paid_amount = Dec(str(max(0, float(order.paid_amount) - refunded)))
    await session.commit()
    await session.refresh(order)
    order_total_after_returns = compute_order_total(order.order_data, exclude_returned=True)
    user_tgid = None
    user_vkid = None
    if order.user_id:
        try:
            async with httpx.AsyncClient() as client:
                r = await client.get(
                    f"{getenv('USERS_SERVICE_URL', 'http://users-service:8001')}/users/{order.user_id}",
                    timeout=5.0
                )
                if r.status_code == 200:
                    ud = r.json()
                    user_tgid = ud.get("tgid")
                    user_vkid = ud.get("vkid")
        except Exception:
            pass
    delivery_info = None
    if order.delivery:
        delivery_info = DeliveryInfoForAdminClass(
            delivery_status_id=order.delivery.delivery_status_id,
            delivery_status_name=order.delivery.delivery_status.name if order.delivery.delivery_status else None,
            additional_info=order.delivery.additional_info
        )
    return AdminOrderResponse(
        id=order.id,
        user_id=order.user_id,
        order_data=await admin_order_data_for_response(session, order_id=order.id, order_data=order.order_data),
        status=order.status,
        created_at=order.created_at,
        updated_at=order.updated_at,
        paid_amount=float(order.paid_amount),
        order_total=order_total_after_returns,
        tracking_number=getattr(order, "tracking_number", None),
        delivery=delivery_info,
        phone_number=order.phone_number,
        is_from_stock=order.is_from_stock,
        cancel_reason=order.cancel_reason,
        refund_on_cancel=order.refund_on_cancel,
        user_tgid=user_tgid,
        user_vkid=user_vkid,
        recipient_name=getattr(order, "recipient_name", None),
    )



